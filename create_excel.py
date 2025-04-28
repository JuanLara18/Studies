import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
import datetime

def create_university_excel(output_file="Information.xlsx"):
    # Crear un libro de Excel
    wb = Workbook()
    
    # Eliminar la hoja por defecto
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Definir estilo para encabezados
    header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # 1. HOJA DE UNIVERSIDADES
    university_data = {
        'Univ_ID': ['UNIV001', 'UNIV002', 'UNIV003'],
        'Country': ['', '', ''],
        'City': ['', '', ''],
        'University': ['', '', ''],
        'Website': ['', '', ''],
        'Type': ['', '', ''],
        'Size': ['', '', ''],
        'Campus Environment': ['', '', ''],
        'Main Language': ['', '', ''],
        'Other Languages': ['', '', ''],
        'Year Established': ['', '', ''],
        'Student Population': ['', '', ''],
        'Faculty-Student Ratio': ['', '', ''],
        'Acceptance Rate (%)': ['', '', ''],
        'Global Ranking (QS)': ['', '', ''],
        'Global Ranking (THE)': ['', '', ''],
        'Subject Ranking': ['', '', ''],
        'Research Expenditure (USD)': ['', '', ''],
        'Endowment (USD)': ['', '', ''],
        'Notable Alumni': ['', '', ''],
        'Official Contact Email': ['', '', ''],
        'Notes': ['', '', '']
    }
    
    df_university = pd.DataFrame(university_data)
    sheet = wb.create_sheet("1_University")
    
    # Aplicar estilos a encabezados
    for r in dataframe_to_rows(df_university, index=False, header=True):
        sheet.append(r)
    
    for col in range(1, len(df_university.columns) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        sheet.column_dimensions[get_column_letter(col)].width = 20
    
    # Configurar validación de datos para campos específicos
    type_validation = DataValidation(type="list", formula1='"Public,Private"')
    sheet.add_data_validation(type_validation)
    type_validation.add(f"F2:F{len(df_university) + 1}")
    
    size_validation = DataValidation(type="list", formula1='"Large,Medium,Small"')
    sheet.add_data_validation(size_validation)
    size_validation.add(f"G2:G{len(df_university) + 1}")
    
    environment_validation = DataValidation(type="list", formula1='"Urban,Suburban,Rural"')
    sheet.add_data_validation(environment_validation)
    environment_validation.add(f"H2:H{len(df_university) + 1}")
    
    # 2. HOJA DE PROGRAMAS
    program_data = {
        'Prog_ID': ['PROG001', 'PROG002', 'PROG003'],
        'Univ_ID': ['', '', ''],
        'Program Name': ['', '', ''],
        'Degree Type': ['', '', ''],
        'Program Website': ['', '', ''],
        'Duration (Years)': ['', '', ''],
        'Mode': ['', '', ''],
        'Number of Credits': ['', '', ''],
        'Tuition Fee (per year)': ['', '', ''],
        'Currency': ['', '', ''],
        'Main Areas of Focus': ['', '', ''],
        'Application Deadline': ['', '', ''],
        'Admission Seasons': ['', '', ''],
        'Start Date': ['', '', ''],
        'Cohort Size': ['', '', ''],
        'Language Requirement': ['', '', ''],
        'Prerequisites': ['', '', ''],
        'Funding Options': ['', '', ''],
        'Program Coordinator': ['', '', ''],
        'Contact Email': ['', '', ''],
        'Notes': ['', '', '']
    }
    
    df_program = pd.DataFrame(program_data)
    sheet = wb.create_sheet("2_Program")
    
    for r in dataframe_to_rows(df_program, index=False, header=True):
        sheet.append(r)
    
    for col in range(1, len(df_program.columns) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        sheet.column_dimensions[get_column_letter(col)].width = 20
    
    # Validaciones para Programas
    degree_validation = DataValidation(type="list", formula1='"Master\'s,Ph.D.,Certificate,Diploma"')
    sheet.add_data_validation(degree_validation)
    degree_validation.add(f"D2:D{len(df_program) + 1}")
    
    mode_validation = DataValidation(type="list", formula1='"Full-time,Part-time,Online,Hybrid"')
    sheet.add_data_validation(mode_validation)
    mode_validation.add(f"G2:G{len(df_program) + 1}")
    
    season_validation = DataValidation(type="list", formula1='"Fall,Spring,Summer,Winter,Multiple"')
    sheet.add_data_validation(season_validation)
    season_validation.add(f"M2:M{len(df_program) + 1}")
    
    # 3. HOJA DE LABORATORIOS E INVESTIGACIÓN
    lab_data = {
        'Lab_ID': ['LAB001', 'LAB002', 'LAB003'],
        'Univ_ID': ['', '', ''],
        'Prog_ID': ['', '', ''],
        'Laboratory / Center Name': ['', '', ''],
        'Department/Faculty': ['', '', ''],
        'Research Fields': ['', '', ''],
        'Website': ['', '', ''],
        'Lab Director': ['', '', ''],
        'Contact Email': ['', '', ''],
        'Key Researchers': ['', '', ''],
        'Location (Building)': ['', '', ''],
        'Number of Active Projects': ['', '', ''],
        'Grant Funding (USD)': ['', '', ''],
        'Industry Collaborations': ['', '', ''],
        'Facilities': ['', '', ''],
        'Annual Publications': ['', '', ''],
        'Student Positions Available': ['', '', ''],
        'Lab Ranking (if available)': ['', '', ''],
        'Notes': ['', '', '']
    }
    
    df_lab = pd.DataFrame(lab_data)
    sheet = wb.create_sheet("3_Lab-Research")
    
    for r in dataframe_to_rows(df_lab, index=False, header=True):
        sheet.append(r)
    
    for col in range(1, len(df_lab.columns) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        sheet.column_dimensions[get_column_letter(col)].width = 20
    
    # 4. HOJA DE BECAS Y FINANCIAMIENTO
    scholarship_data = {
        'Scholarship_ID': ['SCH001', 'SCH002', 'SCH003'],
        'Univ_ID': ['', '', ''],
        'Prog_ID': ['', '', ''],
        'Scholarship Name': ['', '', ''],
        'Type of Funding': ['', '', ''],
        'Amount': ['', '', ''],
        'Currency': ['', '', ''],
        'Eligibility Criteria': ['', '', ''],
        'Competitiveness': ['', '', ''],
        'Number of Awards': ['', '', ''],
        'Application Deadline': ['', '', ''],
        'Notification Date': ['', '', ''],
        'Disbursement Schedule': ['', '', ''],
        'Renewal Conditions': ['', '', ''],
        'Selection Process': ['', '', ''],
        'Scholarship Website': ['', '', ''],
        'Contact Person': ['', '', ''],
        'Contact Email': ['', '', ''],
        'Notes': ['', '', '']
    }
    
    df_scholarship = pd.DataFrame(scholarship_data)
    sheet = wb.create_sheet("4_Scholarships")
    
    for r in dataframe_to_rows(df_scholarship, index=False, header=True):
        sheet.append(r)
    
    for col in range(1, len(df_scholarship.columns) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        sheet.column_dimensions[get_column_letter(col)].width = 20
    
    # Validaciones para Becas
    funding_validation = DataValidation(type="list", formula1='"Full Tuition,Partial Tuition,Living Stipend,Travel Grant,Research Grant,Mixed"')
    sheet.add_data_validation(funding_validation)
    funding_validation.add(f"E2:E{len(df_scholarship) + 1}")
    
    competitiveness_validation = DataValidation(type="list", formula1='"High,Medium,Low"')
    sheet.add_data_validation(competitiveness_validation)
    competitiveness_validation.add(f"I2:I{len(df_scholarship) + 1}")
    
    # 5. HOJA DE REQUISITOS DE ADMISIÓN
    admission_data = {
        'Admission_ID': ['ADM001', 'ADM002', 'ADM003'],
        'Univ_ID': ['', '', ''],
        'Prog_ID': ['', '', ''],
        'Minimum GPA': ['', '', ''],
        'GPA Scale': ['', '', ''],
        'Required Exams': ['', '', ''],
        'Minimum Scores': ['', '', ''],
        'Language Test Validity (years)': ['', '', ''],
        'Letters of Recommendation': ['', '', ''],
        'Statement of Purpose': ['', '', ''],
        'Resume / CV': ['', '', ''],
        'Interview Requirement': ['', '', ''],
        'Research Proposal': ['', '', ''],
        'Experience Required': ['', '', ''],
        'Portfolio/Writing Samples': ['', '', ''],
        'Application Deadline': ['', '', ''],
        'Application Fee (USD)': ['', '', ''],
        'Rolling Admission': ['', '', ''],
        'Other Requirements': ['', '', ''],
        'Notes': ['', '', '']
    }
    
    df_admission = pd.DataFrame(admission_data)
    sheet = wb.create_sheet("5_Admission")
    
    for r in dataframe_to_rows(df_admission, index=False, header=True):
        sheet.append(r)
    
    for col in range(1, len(df_admission.columns) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        sheet.column_dimensions[get_column_letter(col)].width = 20
    
    # Validaciones para Admisión
    yes_no_validation = DataValidation(type="list", formula1='"Yes,No,Optional"')
    sheet.add_data_validation(yes_no_validation)
    yes_no_validation.add(f"K2:K{len(df_admission) + 1}")
    yes_no_validation.add(f"L2:L{len(df_admission) + 1}")
    yes_no_validation.add(f"M2:M{len(df_admission) + 1}")
    yes_no_validation.add(f"R2:R{len(df_admission) + 1}")
    
    # 6. HOJA DE COSTO DE VIDA Y LOGÍSTICA
    cost_data = {
        'Cost_ID': ['CST001', 'CST002', 'CST003'],
        'Univ_ID': ['', '', ''],
        'City': ['', '', ''],
        'Country': ['', '', ''],
        'Currency': ['', '', ''],
        'Estimated Monthly Living Costs': ['', '', ''],
        'Housing Type': ['', '', ''],
        'Housing Costs': ['', '', ''],
        'Food/Groceries': ['', '', ''],
        'Public Transportation': ['', '', ''],
        'Utilities': ['', '', ''],
        'Health Insurance': ['', '', ''],
        'Textbooks & Supplies': ['', '', ''],
        'Climate': ['', '', ''],
        'Safety Rating': ['', '', ''],
        'Part-time Work Opportunities': ['', '', ''],
        'Visa Cost': ['', '', ''],
        'Visa Process': ['', '', ''],
        'Student Services': ['', '', ''],
        'Notes': ['', '', '']
    }
    
    df_cost = pd.DataFrame(cost_data)
    sheet = wb.create_sheet("6_Cost of Living")
    
    for r in dataframe_to_rows(df_cost, index=False, header=True):
        sheet.append(r)
    
    for col in range(1, len(df_cost.columns) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        sheet.column_dimensions[get_column_letter(col)].width = 20
    
    # Validaciones para Costos
    housing_validation = DataValidation(type="list", formula1='"University Dorm,Off-campus Apartment,Shared Apartment,Host Family,Multiple Options"')
    sheet.add_data_validation(housing_validation)
    housing_validation.add(f"G2:G{len(df_cost) + 1}")
    
    safety_validation = DataValidation(type="list", formula1='"Very Safe,Safe,Average,Below Average,Unsafe"')
    sheet.add_data_validation(safety_validation)
    safety_validation.add(f"O2:O{len(df_cost) + 1}")
    
    # 7. HOJA DE OPORTUNIDADES PROFESIONALES Y RESULTADOS
    outcome_data = {
        'Outcome_ID': ['OUT001', 'OUT002', 'OUT003'],
        'Univ_ID': ['', '', ''],
        'Prog_ID': ['', '', ''],
        'Employability Rate (%)': ['', '', ''],
        'Average Starting Salary': ['', '', ''],
        'Currency': ['', '', ''],
        'Time to First Job (months)': ['', '', ''],
        'Top Employers': ['', '', ''],
        'Internship Opportunities': ['', '', ''],
        'Industry Partnerships': ['', '', ''],
        'Alumni Network Size': ['', '', ''],
        'Alumni Events': ['', '', ''],
        'Alumni Mentorship Programs': ['', '', ''],
        'Further Study Rate (%)': ['', '', ''],
        'Job Satisfaction (1-5)': ['', '', ''],
        'Career Support Services': ['', '', ''],
        'Visa Extension Options': ['', '', ''],
        'Notes': ['', '', '']
    }
    
    df_outcome = pd.DataFrame(outcome_data)
    sheet = wb.create_sheet("7_Outcomes")
    
    for r in dataframe_to_rows(df_outcome, index=False, header=True):
        sheet.append(r)
    
    for col in range(1, len(df_outcome.columns) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        sheet.column_dimensions[get_column_letter(col)].width = 20
    
    # 8. HOJA DE NOTAS PERSONALES Y PRIORIZACIÓN
    notes_data = {
        'Notes_ID': ['NOT001', 'NOT002', 'NOT003'],
        'Univ_ID': ['', '', ''],
        'Prog_ID': ['', '', ''],
        'Personal Interest Level': ['', '', ''],
        'Alignment with Career Goals': ['', '', ''],
        'Cultural Fit': ['', '', ''],
        'Family/Friends Nearby': ['', '', ''],
        'Personal Comments': ['', '', ''],
        'Date of Last Review': ['', '', ''],
        'Next Steps': ['', '', ''],
        'Final Decision': ['', '', '']
    }
    
    df_notes = pd.DataFrame(notes_data)
    sheet = wb.create_sheet("8_Notes")
    
    for r in dataframe_to_rows(df_notes, index=False, header=True):
        sheet.append(r)
    
    for col in range(1, len(df_notes.columns) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        sheet.column_dimensions[get_column_letter(col)].width = 20
    
    # Validaciones para Notas
    interest_validation = DataValidation(type="list", formula1='"1-Low,2,3-Medium,4,5-High"')
    sheet.add_data_validation(interest_validation)
    interest_validation.add(f"D2:D{len(df_notes) + 1}")
    
    decision_validation = DataValidation(type="list", formula1='"Shortlist,Backup,Rejected,Top Choice,Applied,Accepted,Declined"')
    sheet.add_data_validation(decision_validation)
    decision_validation.add(f"J2:J{len(df_notes) + 1}")
    
    # 9. HOJA DE CRONOGRAMA DE APLICACIÓN
    timeline_data = {
        'Timeline_ID': ['TL001', 'TL002', 'TL003'],
        'Univ_ID': ['', '', ''],
        'Prog_ID': ['', '', ''],
        'Program Name': ['', '', ''],
        'University': ['', '', ''],
        'Program Deadline': ['', '', ''],
        'Application Start Date': ['', '', ''],
        'Document Preparation': ['', '', ''],
        'Test Date(s)': ['', '', ''],
        'Letter of Rec Deadline': ['', '', ''],
        'Scholarship Deadline': ['', '', ''],
        'Expected Response Date': ['', '', ''],
        'Deposit Due Date': ['', '', ''],
        'Visa Application Date': ['', '', ''],
        'Housing Application': ['', '', ''],
        'Orientation Date': ['', '', ''],
        'Program Start Date': ['', '', ''],
        'Status': ['', '', ''],
        'Priority': ['', '', ''],
        'Notes': ['', '', '']
    }
    
    df_timeline = pd.DataFrame(timeline_data)
    sheet = wb.create_sheet("9_Timeline")
    
    for r in dataframe_to_rows(df_timeline, index=False, header=True):
        sheet.append(r)
    
    for col in range(1, len(df_timeline.columns) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        sheet.column_dimensions[get_column_letter(col)].width = 20
    
    # Validaciones para Timeline
    status_validation = DataValidation(type="list", formula1='"Not Started,In Progress,Completed,Missed,NA"')
    sheet.add_data_validation(status_validation)
    status_validation.add(f"R2:R{len(df_timeline) + 1}")
    
    priority_validation = DataValidation(type="list", formula1='"High,Medium,Low"')
    sheet.add_data_validation(priority_validation)
    priority_validation.add(f"S2:S{len(df_timeline) + 1}")
    
    # 10. HOJA DE DASHBOARD/RESUMEN
    # Esta es una hoja de resumen que podría incluir fórmulas o tablas dinámicas más adelante
    sheet = wb.create_sheet("10_Dashboard")
    sheet['A1'] = "Universidad y Programa Comparaciones - Dashboard"
    sheet['A1'].font = Font(name='Arial', size=14, bold=True)
    
    # Instrucciones básicas para el dashboard
    sheet['A3'] = "Instrucciones:"
    sheet['A4'] = "1. Esta hoja está diseñada para visualizar comparaciones entre programas."
    sheet['A5'] = "2. Puedes agregar tablas dinámicas comparando criterios importantes."
    sheet['A6'] = "3. Recomendado: Agregar gráficas de Ranking vs. Costo vs. Empleabilidad."
    
    # Posibles secciones para el dashboard
    sheet['A8'] = "Programas por Puntuación Personal (agregar tabla dinámica)"
    sheet['A10'] = "Programas por Costo Total (agregar tabla dinámica)"
    sheet['A12'] = "Programas por Oportunidades de Financiamiento (agregar tabla dinámica)"
    sheet['A14'] = "Cronograma Visual de Plazos (agregar gráfica)"
    
    # Guardar el archivo
    wb.save(output_file)
    print(f"Archivo Excel creado exitosamente: {output_file}")
    return output_file

# Ejecutar la función para crear el archivo
if __name__ == "__main__":
    excel_file = create_university_excel()
    print(f"Se ha creado el archivo: {excel_file}")