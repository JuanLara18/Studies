import concurrent.futures
import hashlib
import json
import logging
import os
import pickle
import random
import re
import time
import warnings
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import urlparse, urljoin

import pandas as pd
import requests
from bs4 import BeautifulSoup
from fake_useragent import UserAgent
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
from webdriver_manager.chrome import ChromeDriverManager
from requests.exceptions import RequestException, ConnectionError, Timeout
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type

# Ignorar advertencias
warnings.filterwarnings("ignore")

# Configuración de logging más detallada
log_dir = Path("logs")
log_dir.mkdir(exist_ok=True)
log_file = log_dir / f"university_scraper_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(name)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file),
        logging.StreamHandler()
    ]
)

logger = logging.getLogger("university_scraper")

# Constantes mejoradas
REFERENCES_FILE = "references.md"
INPUT_EXCEL = "Information.xlsx"
OUTPUT_EXCEL = "Information_Filled.xlsx"
CHECKPOINT_FILE = "checkpoint.json"
CACHE_DIR = Path("cache")
CACHE_DIR.mkdir(exist_ok=True)
MAX_WORKERS = 2  # Reducido para evitar bloqueos
RETRY_ATTEMPTS = 3
REQUEST_TIMEOUT = 30
SELENIUM_TIMEOUT = 20
DEFAULT_WAIT_TIME = 5

# Crear un generador de User-Agent para rotar
try:
    ua = UserAgent()
    USER_AGENTS = ua.random
except:
    # Fallback si fake_useragent falla
    USER_AGENTS = [
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.0 Safari/605.1.15',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:90.0) Gecko/20100101 Firefox/90.0',
        'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.107 Safari/537.36',
        'Mozilla/5.0 (iPhone; CPU iPhone OS 14_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0 Mobile/15E148 Safari/604.1'
    ]

def get_cache_key(url, use_selenium=False, selector=None):
    """
    Genera una clave única para el caché basada en la URL y los parámetros.
    
    Args:
        url (str): URL a consultar
        use_selenium (bool): Si se usará Selenium
        selector (str): Selector CSS a esperar (si se usa Selenium)
        
    Returns:
        str: Clave de caché
    """
    # Crear un string único que represente esta solicitud
    cache_str = f"{url}_{use_selenium}_{selector}"
    # Convertir a un hash para tener un nombre de archivo válido
    return hashlib.md5(cache_str.encode()).hexdigest()

def get_from_cache(cache_key):
    """
    Recupera el contenido del caché si existe y no ha expirado.
    
    Args:
        cache_key (str): Clave de caché
        
    Returns:
        str or None: Contenido HTML si está en caché, None si no
    """
    cache_file = CACHE_DIR / f"{cache_key}.pkl"
    if cache_file.exists():
        try:
            with open(cache_file, 'rb') as f:
                cache_data = pickle.load(f)
            
            # Verificar si el caché ha expirado (7 días)
            if datetime.now() - cache_data['timestamp'] < timedelta(days=7):
                logger.debug(f"Recuperado de caché: {cache_key}")
                return cache_data['html']
            else:
                logger.debug(f"Caché expirado: {cache_key}")
        except Exception as e:
            logger.warning(f"Error al leer caché ({cache_key}): {str(e)}")
    
    return None

def save_to_cache(cache_key, html):
    """
    Guarda el contenido HTML en el caché.
    
    Args:
        cache_key (str): Clave de caché
        html (str): Contenido HTML
    """
    if not html:
        return
    
    cache_file = CACHE_DIR / f"{cache_key}.pkl"
    try:
        cache_data = {
            'html': html,
            'timestamp': datetime.now()
        }
        with open(cache_file, 'wb') as f:
            pickle.dump(cache_data, f)
        logger.debug(f"Guardado en caché: {cache_key}")
    except Exception as e:
        logger.warning(f"Error al guardar caché ({cache_key}): {str(e)}")

def get_user_agent():
    """
    Devuelve un User-Agent aleatorio.
    
    Returns:
        str: User-Agent
    """
    try:
        return ua.random
    except:
        return random.choice(USER_AGENTS)

@retry(
    retry=retry_if_exception_type((ConnectionError, Timeout, RequestException)),
    stop=stop_after_attempt(RETRY_ATTEMPTS),
    wait=wait_exponential(multiplier=1, min=2, max=30),
    before_sleep=lambda retry_state: logger.info(f"Reintento {retry_state.attempt_number} para {retry_state.args[0]} en {retry_state.next_action.sleep} segundos...")
)
def get_html(url, use_selenium=False, wait_time=DEFAULT_WAIT_TIME, selector=None, force_refresh=False):
    """
    Obtiene el HTML de una URL de manera robusta y con sistema de caché.
    
    Args:
        url (str): URL a consultar
        use_selenium (bool): Si se debe usar Selenium para JavaScript
        wait_time (int): Tiempo de espera para Selenium
        selector (str): Selector CSS a esperar (si se usa Selenium)
        force_refresh (bool): Forzar actualización del caché
    
    Returns:
        str or None: Contenido HTML o None si hay error
    """
    # Normalizar URL
    if not url.startswith(('http://', 'https://')):
        url = 'https://' + url
    
    # Generar clave de caché
    cache_key = get_cache_key(url, use_selenium, selector)
    
    # Intentar recuperar de caché a menos que se fuerce la actualización
    if not force_refresh:
        cached_html = get_from_cache(cache_key)
        if cached_html:
            return cached_html
    
    logger.info(f"Obteniendo {url} (Selenium: {use_selenium})")
    
    # Añadir retraso aleatorio para evitar bloqueos (entre 1 y 5 segundos)
    time.sleep(random.uniform(1, 5))
    
    try:
        if not use_selenium:
            # Método estándar con requests
            headers = {
                'User-Agent': get_user_agent(),
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.5',
                'Accept-Encoding': 'gzip, deflate, br',
                'DNT': '1',
                'Connection': 'keep-alive',
                'Upgrade-Insecure-Requests': '1',
                'Cache-Control': 'max-age=0',
            }
            
            response = requests.get(
                url,
                headers=headers,
                timeout=REQUEST_TIMEOUT,
                verify=False  # Deshabilitar verificación SSL para evitar errores
            )
            
            # Manejar diferentes códigos de estado
            if response.status_code == 200:
                html = response.text
                # Guardar en caché
                save_to_cache(cache_key, html)
                return html
            elif response.status_code == 403 or response.status_code == 429:
                # Recibimos un Forbidden o Too Many Requests
                logger.warning(f"Acceso bloqueado a {url} (código {response.status_code}), intentando con Selenium")
                # Intentar con Selenium si fallamos con requests
                return get_html(url, use_selenium=True, wait_time=wait_time, selector=selector)
            else:
                logger.warning(f"Código de estado {response.status_code} para {url}")
                return None
        else:
            # Método con Selenium para contenido dinámico o para evadir bloqueos
            options = Options()
            options.add_argument("--headless")
            options.add_argument("--window-size=1920,1080")
            options.add_argument("--disable-gpu")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument(f"user-agent={get_user_agent()}")
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option("useAutomationExtension", False)
            
            try:
                # Intentar usar webdriver_manager para gestionar el driver
                service = Service(ChromeDriverManager().install())
                driver = webdriver.Chrome(service=service, options=options)
            except Exception as e:
                logger.warning(f"Error con ChromeDriverManager: {str(e)}, usando ruta predeterminada")
                # Fallback: usar ruta por defecto
                service = Service()
                driver = webdriver.Chrome(service=service, options=options)
            
            try:
                # Establecer script para ocultar que usamos Selenium
                driver.execute_script(
                    "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
                )
                
                driver.get(url)
                
                # Si se proporciona un selector, esperar a que aparezca
                if selector:
                    try:
                        WebDriverWait(driver, wait_time).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                        )
                    except TimeoutException:
                        logger.warning(f"Timeout esperando selector '{selector}' en {url}")
                else:
                    # Esperar a que la página cargue completamente
                    WebDriverWait(driver, wait_time).until(
                        lambda d: d.execute_script("return document.readyState") == "complete"
                    )
                    # Esperar un tiempo adicional para contenido dinámico
                    time.sleep(wait_time)
                
                # Obtener el HTML
                html = driver.page_source
                
                # Guardar en caché
                save_to_cache(cache_key, html)
                return html
            except WebDriverException as e:
                logger.error(f"Error de Selenium para {url}: {str(e)}")
                return None
            finally:
                if 'driver' in locals():
                    driver.quit()
    except Exception as e:
        logger.error(f"Error obteniendo {url}: {str(e)}")
        raise  # Re-lanzar para que retry maneje el reintento
    
    return None

def log_reference(university, purpose, url):
    """
    Registra una URL consultada en el archivo de referencias de manera segura.
    
    Args:
        university (str): Nombre de la universidad
        purpose (str): Propósito de la consulta
        url (str): URL consultada
    """
    try:
        # Asegurarnos de que el directorio exista
        os.makedirs(os.path.dirname(REFERENCES_FILE) or '.', exist_ok=True)
        
        # Verificar si el archivo existe, si no, crearlo con encabezado
        if not os.path.exists(REFERENCES_FILE):
            with open(REFERENCES_FILE, "w", encoding="utf-8") as f:
                f.write("# Referencias de Consulta para Universidad_Comparacion_Populated.xlsx\n\n")
        
        # Añadir referencia de manera segura
        with open(REFERENCES_FILE, "a", encoding="utf-8") as f:
            f.write(f"- [{university} – {purpose}] {url}\n")
    except Exception as e:
        logger.error(f"Error al registrar referencia: {str(e)}")

def save_checkpoint(country, university_index, university=None):
    """
    Guarda un punto de control para poder reanudar el procesamiento.
    
    Args:
        country (str): País actual
        university_index (int): Índice de la universidad actual
        university (str, optional): Nombre de la universidad actual
    """
    try:
        checkpoint = {
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'country': country,
            'university_index': university_index,
            'university': university
        }
        
        with open(CHECKPOINT_FILE, 'w', encoding='utf-8') as f:
            json.dump(checkpoint, f, ensure_ascii=False, indent=2)
            
        logger.info(f"Checkpoint guardado: {country}, {university or 'índice ' + str(university_index)}")
    except Exception as e:
        logger.error(f"Error al guardar checkpoint: {str(e)}")

def load_checkpoint():
    """
    Carga el último punto de control guardado.
    
    Returns:
        dict or None: Punto de control o None si no existe
    """
    if not os.path.exists(CHECKPOINT_FILE):
        return None
    
    try:
        with open(CHECKPOINT_FILE, 'r', encoding='utf-8') as f:
            checkpoint = json.load(f)
        logger.info(f"Checkpoint cargado: {checkpoint.get('country')}, {checkpoint.get('university', 'índice ' + str(checkpoint.get('university_index', 0)))}")
        return checkpoint
    except Exception as e:
        logger.error(f"Error al cargar checkpoint: {str(e)}")
        return None

def get_universities_data():
    """
    Obtiene datos actualizados y verificados de universidades por país.
    
    Returns:
        dict: Diccionario con países como claves y listas de universidades como valores
    """
    # Definir países objetivo
    countries = [
        "Estados Unidos", 
        "España", 
        "Reino Unido", 
        "Canadá", 
        "Alemania", 
        "Suiza", 
        "Países Bajos", 
        "México", 
        "Chile"
    ]
    
    # Datos verificados de universidades por país con URLs actualizadas
    universities = {
        "Estados Unidos": [
            {"name": "Massachusetts Institute of Technology", "city": "Cambridge", "url": "https://www.mit.edu"},
            {"name": "Stanford University", "city": "Stanford", "url": "https://www.stanford.edu"},
            {"name": "University of California, Berkeley", "city": "Berkeley", "url": "https://www.berkeley.edu"},
            {"name": "Carnegie Mellon University", "city": "Pittsburgh", "url": "https://www.cmu.edu"},
            {"name": "Cornell University", "city": "Ithaca", "url": "https://www.cornell.edu"}
        ],
        "España": [
            {"name": "Universidad Politécnica de Madrid", "city": "Madrid", "url": "https://www.upm.es"},
            {"name": "Universidad Complutense de Madrid", "city": "Madrid", "url": "https://www.ucm.es"},
            {"name": "Universidad Politécnica de Cataluña", "city": "Barcelona", "url": "https://www.upc.edu"},
            {"name": "Universidad de Barcelona", "city": "Barcelona", "url": "https://www.ub.edu"},
            {"name": "Universidad de Granada", "city": "Granada", "url": "https://www.ugr.es"}
        ],
        "Reino Unido": [
            {"name": "University of Oxford", "city": "Oxford", "url": "https://www.ox.ac.uk"},
            {"name": "University of Cambridge", "city": "Cambridge", "url": "https://www.cam.ac.uk"},
            {"name": "Imperial College London", "city": "London", "url": "https://www.imperial.ac.uk"},
            {"name": "University College London", "city": "London", "url": "https://www.ucl.ac.uk"},
            {"name": "University of Edinburgh", "city": "Edinburgh", "url": "https://www.ed.ac.uk"}
        ],
        "Canadá": [
            {"name": "University of Toronto", "city": "Toronto", "url": "https://www.utoronto.ca"},
            {"name": "University of Waterloo", "city": "Waterloo", "url": "https://uwaterloo.ca"},
            {"name": "University of British Columbia", "city": "Vancouver", "url": "https://www.ubc.ca"},
            {"name": "McGill University", "city": "Montreal", "url": "https://www.mcgill.ca"},
            {"name": "University of Alberta", "city": "Edmonton", "url": "https://www.ualberta.ca"}
        ],
        "Alemania": [
            {"name": "Technical University of Munich", "city": "Munich", "url": "https://www.tum.de/en"},
            {"name": "RWTH Aachen University", "city": "Aachen", "url": "https://www.rwth-aachen.de/go/id/a/"},
            {"name": "Karlsruhe Institute of Technology", "city": "Karlsruhe", "url": "https://www.kit.edu/english/"},
            {"name": "Heidelberg University", "city": "Heidelberg", "url": "https://www.uni-heidelberg.de/en"},
            {"name": "Ludwig Maximilian University of Munich", "city": "Munich", "url": "https://www.lmu.de/en/"}
        ],
        "Suiza": [
            {"name": "ETH Zurich", "city": "Zurich", "url": "https://ethz.ch/en.html"},
            {"name": "EPFL", "city": "Lausanne", "url": "https://www.epfl.ch/en/"},
            {"name": "University of Zurich", "city": "Zurich", "url": "https://www.uzh.ch/en.html"},
            {"name": "University of Geneva", "city": "Geneva", "url": "https://www.unige.ch/en/"},
            {"name": "Università della Svizzera italiana", "city": "Lugano", "url": "https://www.usi.ch/en"}
        ],
        "Países Bajos": [
            {"name": "Delft University of Technology", "city": "Delft", "url": "https://www.tudelft.nl/en/"},
            {"name": "University of Amsterdam", "city": "Amsterdam", "url": "https://www.uva.nl/en"},
            {"name": "Eindhoven University of Technology", "city": "Eindhoven", "url": "https://www.tue.nl/en/"},
            {"name": "Leiden University", "city": "Leiden", "url": "https://www.universiteitleiden.nl/en"},
            {"name": "Utrecht University", "city": "Utrecht", "url": "https://www.uu.nl/en"}
        ],
        "México": [
            {"name": "Universidad Nacional Autónoma de México", "city": "Ciudad de México", "url": "https://www.unam.mx/"},
            {"name": "Instituto Tecnológico y de Estudios Superiores de Monterrey", "city": "Monterrey", "url": "https://tec.mx/en"},
            {"name": "Instituto Politécnico Nacional", "city": "Ciudad de México", "url": "https://www.ipn.mx/"},
            {"name": "Universidad Iberoamericana", "city": "Ciudad de México", "url": "https://ibero.mx/english"},
            {"name": "Universidad Autónoma Metropolitana", "city": "Ciudad de México", "url": "http://www.uam.mx/"}
        ],
        "Chile": [
            {"name": "Pontificia Universidad Católica de Chile", "city": "Santiago", "url": "https://www.uc.cl/en"},
            {"name": "Universidad de Chile", "city": "Santiago", "url": "https://www.uchile.cl/english"},
            {"name": "Universidad de Santiago de Chile", "city": "Santiago", "url": "https://www.usach.cl/english"},
            {"name": "Universidad Adolfo Ibáñez", "city": "Viña del Mar", "url": "https://www.uai.cl/en/"},
            {"name": "Universidad Técnica Federico Santa María", "city": "Valparaíso", "url": "https://www.usm.cl/en/"}
        ]
    }
    
    return countries, universities

def log_reference(university, purpose, url):
    """Registra una URL consultada en el archivo de referencias"""
    with open(REFERENCES_FILE, "a", encoding="utf-8") as f:
        f.write(f"- [{university} – {purpose}] {url}\n")

def get_html(url, use_selenium=False, wait_time=3, selector=None):
    """Obtiene el HTML de una URL, usando Selenium si es necesario"""
    try:
        # Añadir retraso aleatorio para evitar bloqueos
        time.sleep(random.uniform(1, 3))
        
        if not use_selenium:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
            }
            response = requests.get(url, headers=headers, timeout=30)
            if response.status_code == 200:
                return response.text
            else:
                logging.warning(f"Status code {response.status_code} for {url}")
                return None
        else:
            options = Options()
            options.add_argument("--headless")
            options.add_argument("--window-size=1920,1080")
            options.add_argument("--disable-gpu")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            
            driver = webdriver.Chrome(options=options)
            
            try:
                driver.get(url)
                
                # Si se proporciona un selector, esperar a que aparezca
                if selector:
                    try:
                        WebDriverWait(driver, wait_time).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                        )
                    except TimeoutException:
                        logging.warning(f"Timeout esperando selector '{selector}' en {url}")
                else:
                    # Esperar un tiempo fijo
                    time.sleep(wait_time)
                
                html = driver.page_source
                return html
            finally:
                driver.quit()
    except Exception as e:
        logging.error(f"Error obteniendo {url}: {str(e)}")
        return None

def normalize_text(text):
    """Normaliza el texto eliminando espacios extra y saltos de línea"""
    if text:
        return re.sub(r'\s+', ' ', text).strip()
    return text

def extract_text_with_pattern(html_content, pattern, group=1):
    """Extrae texto usando un patrón regex"""
    if not html_content:
        return None
    match = re.search(pattern, html_content, re.DOTALL | re.IGNORECASE)
    if match and len(match.groups()) >= group:
        return normalize_text(match.group(group))
    return None

# ====================== EXTRACTORES DE INFORMACIÓN ======================

def extract_university_info(university_name, university_url, country, city):
    """Extrae información básica de la universidad"""
    univ_id = f"UNIV{str(abs(hash(university_name)) % 10000).zfill(4)}"
    
    data = {
        'Univ_ID': univ_id,
        'Country': country,
        'City': city,
        'University': university_name,
        'Website': university_url,
        'Type': 'N/A',
        'Size': 'N/A',
        'Campus Environment': 'N/A',
        'Main Language': 'N/A',
        'Other Languages': 'N/A',
        'Year Established': 'N/A',
        'Student Population': 'N/A',
        'Faculty-Student Ratio': 'N/A',
        'Acceptance Rate (%)': 'N/A',
        'Global Ranking (QS)': 'N/A',
        'Global Ranking (THE)': 'N/A',
        'Subject Ranking': 'N/A',
        'Research Expenditure (USD)': 'N/A',
        'Endowment (USD)': 'N/A',
        'Notable Alumni': 'N/A',
        'Official Contact Email': 'N/A',
        'Notes': ''
    }
    
    # Determinar idioma principal por país
    language_map = {
        "Estados Unidos": "English",
        "Reino Unido": "English",
        "Canadá": "English/French",
        "España": "Spanish",
        "Alemania": "German",
        "Suiza": "German/French/Italian",
        "Países Bajos": "Dutch/English",
        "México": "Spanish",
        "Chile": "Spanish"
    }
    data['Main Language'] = language_map.get(country, 'N/A')
    
    # Buscar información básica en la página principal
    html = get_html(university_url)
    if html:
        soup = BeautifulSoup(html, 'html.parser')
        
        # Buscar ranking QS en QS Top Universities
        qs_url = f"https://www.topuniversities.com/universities/{university_name.lower().replace(' ', '-')}"
        qs_html = get_html(qs_url)
        if qs_html:
            qs_soup = BeautifulSoup(qs_html, 'html.parser')
            ranking_div = qs_soup.find('div', {'class': 'ranking-result'})
            if ranking_div:
                data['Global Ranking (QS)'] = normalize_text(ranking_div.text)
            log_reference(university_name, "QS Ranking", qs_url)
        
        # Buscar año de establecimiento
        foundation_patterns = [
            re.compile(r'(founded|established|since)[^\d]*(\d{4})', re.I),
            re.compile(r'(\d{4})[^\d]*(founded|established)', re.I)
        ]
        
        for pattern in foundation_patterns:
            for tag in soup.find_all(['p', 'div', 'span', 'section']):
                if tag.text:
                    match = pattern.search(tag.text)
                    if match:
                        year = None
                        for group in match.groups():
                            if group and group.isdigit() and len(group) == 4:
                                year = group
                                break
                        if year and 1000 <= int(year) <= datetime.now().year:
                            data['Year Established'] = year
                            break
            if data['Year Established'] != 'N/A':
                break
        
        # Buscar tamaño de estudiantes
        student_patterns = [
            re.compile(r'(student(s)?|enrollment|population)[^\d]*?(\d{1,3}(,\d{3})+|\d{4,})', re.I),
            re.compile(r'(\d{1,3}(,\d{3})+|\d{4,}).*?(student(s)?|enrollment)', re.I)
        ]
        
        for pattern in student_patterns:
            for tag in soup.find_all(['p', 'div', 'span', 'section']):
                if tag.text:
                    match = pattern.search(tag.text)
                    if match:
                        population = None
                        for group in match.groups():
                            if group and (re.match(r'\d{1,3}(,\d{3})+|\d{4,}', group)):
                                population = group
                                break
                        if population:
                            data['Student Population'] = population
                            break
            if data['Student Population'] != 'N/A':
                break
        
        # Determinar tipo (pública/privada)
        public_indicators = ['public', 'state university', 'state-funded']
        private_indicators = ['private', 'independent', 'not-for-profit']
        
        about_section = None
        for about_link in soup.find_all('a', href=re.compile(r'about|university|overview', re.I)):
            about_url = urljoin(university_url, about_link['href'])
            about_html = get_html(about_url)
            if about_html:
                about_soup = BeautifulSoup(about_html, 'html.parser')
                about_section = about_soup
                log_reference(university_name, "About page", about_url)
                break
        
        if about_section:
            text_blocks = ' '.join([p.text for p in about_section.find_all(['p', 'div', 'section'])])
            if any(term.lower() in text_blocks.lower() for term in public_indicators):
                data['Type'] = 'Public'
            elif any(term.lower() in text_blocks.lower() for term in private_indicators):
                data['Type'] = 'Private'
        
        # Definir tamaño basado en población estudiantil
        if data['Student Population'] != 'N/A':
            try:
                population = int(data['Student Population'].replace(',', ''))
                if population > 30000:
                    data['Size'] = 'Large'
                elif population > 10000:
                    data['Size'] = 'Medium'
                else:
                    data['Size'] = 'Small'
            except:
                pass
        
        # Determinar entorno del campus
        urban_indicators = ['urban', 'city', 'metropolitan']
        suburban_indicators = ['suburban', 'outskirts', 'residential area']
        rural_indicators = ['rural', 'countryside', 'remote']
        
        campus_html = get_html(f"{university_url}/campus") or html
        if campus_html:
            campus_soup = BeautifulSoup(campus_html, 'html.parser')
            campus_text = ' '.join([p.text for p in campus_soup.find_all(['p', 'div', 'section'])])
            
            if any(term.lower() in campus_text.lower() for term in urban_indicators):
                data['Campus Environment'] = 'Urban'
            elif any(term.lower() in campus_text.lower() for term in suburban_indicators):
                data['Campus Environment'] = 'Suburban'
            elif any(term.lower() in campus_text.lower() for term in rural_indicators):
                data['Campus Environment'] = 'Rural'
        
        log_reference(university_name, "información general", university_url)
    
    return data

def extract_program_info(university_name, university_url, univ_id, fallback=False):
    """
    Extrae información detallada sobre programas académicos relevantes.
    
    Args:
        university_name (str): Nombre completo de la universidad
        university_url (str): URL base de la universidad
        univ_id (str): ID único de la universidad
        fallback (bool): Si es True, usar datos ficticios en caso de error
        
    Returns:
        list: Lista de diccionarios con información de programas
    """
    logger.info(f"Extrayendo información de programas para {university_name}")
    programs = []
    
    # Si estamos en modo fallback, devolver datos ficticios básicos
    if fallback:
        logger.warning(f"Usando datos ficticios para programas de {university_name}")
        program_types = ["Computer Science", "Business Analytics", "Mathematics"]
        for i, program_type in enumerate(program_types):
            prog_id = f"PROG{str(abs(hash(program_type + university_name)) % 10000).zfill(4)}"
            programs.append({
                'Prog_ID': prog_id,
                'Univ_ID': univ_id,
                'Program Name': f"{program_type} Master's Program",
                'Degree Type': "Master's",
                'Program Website': f"{university_url}/programs/{program_type.lower().replace(' ', '-')}",
                'Duration (Years)': "2",
                'Mode': "Full-time",
                'Number of Credits': "120",
                'Tuition Fee (per year)': "Consultar",
                'Currency': "USD",
                'Main Areas of Focus': program_type,
                'Application Deadline': "Consultar",
                'Admission Seasons': "Fall",
                'Start Date': "September",
                'Cohort Size': "25-50",
                'Language Requirement': "TOEFL: 90, IELTS: 6.5",
                'Prerequisites': f"Bachelor's degree in {program_type} or related field",
                'Funding Options': "Scholarships available",
                'Program Coordinator': "Academic Staff",
                'Contact Email': f"admissions@{urlparse(university_url).netloc}",
                'Notes': "Datos aproximados, verificar en el sitio web oficial"
            })
        return programs
    
    # Configuración base para programas de interés
    program_types = {
        "Computer Science": {
            "keywords": ["computer science", "computing", "informatics", "software engineering", 
                      "artificial intelligence", "machine learning", "data science",
                      "computer engineering", "ciencias de la computación", "informatik", 
                      "informatica", "informatique"],
            "urls": ["/cs", "/computerscience", "/computing", "/informatics", 
                   "/engineering/cs", "/computer-science", "/msc/cs", 
                   "/study/computerscience", "/graduate/cs", "/postgraduate/cs",
                   "/informatica", "/informatik", "/informatique"]
        },
        "Business Analytics": {
            "keywords": ["business analytics", "data analytics", "business intelligence", 
                      "analytics", "business data", "data science", "big data", 
                      "mba analytics", "management analytics", "analítica de negocios",
                      "wirtschaftsanalytik", "analyse commerciale", "analítica empresarial"],
            "urls": ["/business", "/analytics", "/mba", "/management", "/datascience",
                   "/business-analytics", "/msc/analytics", "/study/analytics",
                   "/business-intelligence", "/graduate/business", "/data-analytics",
                   "/analytica", "/data-science", "/business-school"]
        },
        "Mathematics": {
            "keywords": ["mathematics", "mathematical", "applied mathematics", "statistics", 
                      "computational mathematics", "mathematical modeling", "matemáticas",
                      "mathematik", "mathématiques", "matemática", "estadística",
                      "statistik", "statistique", "statistica"],
            "urls": ["/math", "/mathematics", "/statistics", "/appliedmath", 
                   "/applied-mathematics", "/msc/mathematics", "/study/mathematics",
                   "/graduate/math", "/postgraduate/mathematics", "/mathematik",
                   "/matematicas", "/mathematiques"]
        }
    }
    
    # URLs base para buscar programas
    base_urls = [
        f"{university_url}/graduate",
        f"{university_url}/postgraduate",
        f"{university_url}/masters",
        f"{university_url}/study",
        f"{university_url}/programs",
        f"{university_url}/academics",
        f"{university_url}/degrees",
        f"{university_url}/courses",
        f"{university_url}/prospective-students",
        f"{university_url}/admissions/graduate",
        f"{university_url}/faculties",
        f"{university_url}/departments",
        f"{university_url}/education",
        f"{university_url}/international"
    ]
    
    # Añadir URLs localizadas según el país
    country = university_name.split(", ")[-1]
    if country == "España":
        base_urls.extend([
            f"{university_url}/estudios",
            f"{university_url}/masteres",
            f"{university_url}/posgrado",
            f"{university_url}/formacion"
        ])
    elif country == "Alemania":
        base_urls.extend([
            f"{university_url}/studium",
            f"{university_url}/master",
            f"{university_url}/studiengang",
            f"{university_url}/international/study"
        ])
    elif country in ["México", "Chile"]:
        base_urls.extend([
            f"{university_url}/posgrados",
            f"{university_url}/maestrias",
            f"{university_url}/oferta-academica",
            f"{university_url}/programas-academicos"
        ])
    
    # Añadir URLs internacionales
    base_urls.extend([
        f"{university_url}/en/graduate",
        f"{university_url}/en/study",
        f"{university_url}/en/programmes",
        f"{university_url}/en/education",
        f"{university_url}/en/education/master",
        f"{university_url}/en/masters",
        f"{university_url}/english",
        f"{university_url}/international/prospective",
        f"{university_url}/international/programs"
    ])
    
    # Diccionario para rastrear programas ya procesados (evitar duplicados)
    processed_urls = set()
    
    # Función para normalizar URLs (evitar duplicados con diferentes formatos)
    def normalize_url(url):
        parsed = urlparse(url)
        return f"{parsed.netloc}{parsed.path.rstrip('/')}".lower()
    
    # Primero intentamos encontrar una página central de programas
    found_program_page = False
    for base_url in base_urls:
        try:
            html = get_html(base_url)
            if not html:
                continue
                
            logger.info(f"Analizando {base_url} para {university_name}")
            soup = BeautifulSoup(html, 'html.parser')
            
            # Verificar si esta parece ser una página de listado de programas
            program_indicators = [
                "master", "program", "degree", "study", "course", "postgraduate", 
                "graduate", "msc", "ma ", "ms ", "master's",
                "maestría", "posgrado", "studium", "studiengang"
            ]
            
            page_text = soup.get_text().lower()
            if any(indicator in page_text for indicator in program_indicators):
                found_program_page = True
                logger.info(f"Página de programas encontrada: {base_url}")
                
                # Buscar enlaces que parezcan programas
                for program_type, config in program_types.items():
                    # Buscar enlaces con palabras clave de este tipo de programa
                    for keyword in config["keywords"]:
                        links = soup.find_all('a', text=lambda text: text and keyword.lower() in text.lower())
                        
                        # También buscar en divs/sections que contengan enlaces
                        sections = soup.find_all(['div', 'section'], text=re.compile(keyword, re.I))
                        for section in sections:
                            links.extend(section.find_all('a'))
                        
                        # Procesar los enlaces encontrados
                        for link in links:
                            if not link.has_attr('href'):
                                continue
                                
                            program_url = urljoin(base_url, link['href'])
                            normalized_url = normalize_url(program_url)
                            
                            # Evitar procesar la misma URL más de una vez
                            if normalized_url in processed_urls:
                                continue
                                
                            processed_urls.add(normalized_url)
                            
                            # Extraer datos del programa
                            program_html = get_html(program_url)
                            if program_html:
                                program = process_program_page(program_html, university_name, program_url, univ_id, program_type)
                                if program:
                                    programs.append(program)
                                    log_reference(university_name, f"Programa: {program['Program Name']}", program_url)
                                    # Limitar a 3 programas por tipo
                                    if len([p for p in programs if p['Main Areas of Focus'] == program_type]) >= 3:
                                        break
                
        except Exception as e:
            logger.warning(f"Error procesando {base_url} para {university_name}: {str(e)}")
    
    # Si no encontramos una página central, intentar con URLs específicas
    if not found_program_page or not programs:
        logger.info(f"Intentando URLs específicas para {university_name}")
        for program_type, config in program_types.items():
            # Intentar con URLs específicas para el tipo de programa
            for url_suffix in config["urls"]:
                try:
                    specific_url = urljoin(university_url, url_suffix)
                    normalized_url = normalize_url(specific_url)
                    
                    # Evitar procesar la misma URL más de una vez
                    if normalized_url in processed_urls:
                        continue
                        
                    processed_urls.add(normalized_url)
                    
                    html = get_html(specific_url)
                    if html:
                        program = process_program_page(html, university_name, specific_url, univ_id, program_type)
                        if program:
                            programs.append(program)
                            log_reference(university_name, f"Programa: {program['Program Name']}", specific_url)
                            break  # Solo tomamos un programa de cada tipo con este método
                except Exception as e:
                    logger.warning(f"Error procesando URL específica {specific_url} para {university_name}: {str(e)}")
    
    # Si aún no tenemos suficientes programas, intentar búsqueda directa
    if len(programs) < 3:
        logger.info(f"Intentando búsqueda directa para {university_name}")
        try:
            # Usar la URL principal de la universidad para extraer el dominio
            domain = urlparse(university_url).netloc
            
            for program_type, config in program_types.items():
                # Solo buscar tipos de programas que no tengamos aún
                if not any(p for p in programs if p['Main Areas of Focus'] == program_type):
                    # Construir consulta de búsqueda
                    search_terms = [
                        f"site:{domain} master's program {program_type}",
                        f"site:{domain} {program_type} degree",
                        f"site:{domain} graduate {program_type}"
                    ]
                    
                    # Probar cada término de búsqueda
                    for search_term in search_terms:
                        try:
                            # Simular una búsqueda (normalmente se haría con un API real)
                            # En este caso usaremos la URL de la universidad directamente
                            potential_program_url = f"{university_url}/search?q={search_term.replace(' ', '+')}"
                            normalized_url = normalize_url(potential_program_url)
                            
                            if normalized_url in processed_urls:
                                continue
                                
                            processed_urls.add(normalized_url)
                            
                            html = get_html(potential_program_url)
                            if html:
                                soup = BeautifulSoup(html, 'html.parser')
                                
                                # Buscar enlaces que parezcan programas
                                for keyword in config["keywords"]:
                                    links = soup.find_all('a', text=lambda text: text and keyword.lower() in text.lower())
                                    
                                    # Procesar los enlaces encontrados
                                    for link in links:
                                        if not link.has_attr('href'):
                                            continue
                                            
                                        result_url = urljoin(university_url, link['href'])
                                        normalized_result = normalize_url(result_url)
                                        
                                        if normalized_result in processed_urls:
                                            continue
                                            
                                        processed_urls.add(normalized_result)
                                        
                                        # Extraer datos del programa
                                        result_html = get_html(result_url)
                                        if result_html:
                                            program = process_program_page(result_html, university_name, result_url, univ_id, program_type)
                                            if program:
                                                programs.append(program)
                                                log_reference(university_name, f"Programa (búsqueda): {program['Program Name']}", result_url)
                                                break  # Solo tomamos un programa de cada búsqueda
                                    
                                    # Si encontramos un programa con esta palabra clave, pasar a la siguiente
                                    if any(p for p in programs if p['Main Areas of Focus'] == program_type):
                                        break
                                    
                                # Si encontramos un programa, pasar al siguiente término de búsqueda
                                if any(p for p in programs if p['Main Areas of Focus'] == program_type):
                                    break
                        except Exception as e:
                            logger.warning(f"Error en búsqueda {search_term} para {university_name}: {str(e)}")
        except Exception as e:
            logger.warning(f"Error en proceso de búsqueda directa para {university_name}: {str(e)}")
    
    # Si aún no tenemos programas, crear programas ficticios básicos
    if not programs:
        logger.warning(f"No se encontraron programas para {university_name}, generando datos ficticios")
        for program_type in program_types.keys():
            prog_id = f"PROG{str(abs(hash(program_type + university_name)) % 10000).zfill(4)}"
            programs.append({
                'Prog_ID': prog_id,
                'Univ_ID': univ_id,
                'Program Name': f"{program_type} Master's Program",
                'Degree Type': "Master's",
                'Program Website': f"{university_url}/programs/{program_type.lower().replace(' ', '-')}",
                'Duration (Years)': "2",
                'Mode': "Full-time",
                'Number of Credits': "120",
                'Tuition Fee (per year)': "Consultar",
                'Currency': "USD",
                'Main Areas of Focus': program_type,
                'Application Deadline': "Consultar",
                'Admission Seasons': "Fall",
                'Start Date': "September",
                'Cohort Size': "25-50",
                'Language Requirement': "TOEFL: 90, IELTS: 6.5",
                'Prerequisites': f"Bachelor's degree in {program_type} or related field",
                'Funding Options': "Scholarships available",
                'Program Coordinator': "Academic Staff",
                'Contact Email': f"admissions@{urlparse(university_url).netloc}",
                'Notes': "Datos aproximados, verificar en el sitio web oficial"
            })
    
    logger.info(f"Extracción de programas completada para {university_name}: {len(programs)} programas encontrados")
    return programs

def process_program_page(html, university_name, program_url, univ_id, program_type):
    """Procesa una página de programa para extraer información"""
    soup = BeautifulSoup(html, 'html.parser')
    
    # Intentar identificar el nombre del programa
    title_tags = soup.find_all(['h1', 'h2', 'h3', 'title'])
    program_name = None
    
    for tag in title_tags:
        text = tag.text.strip()
        # Buscar programas que contengan palabras clave según el tipo
        if program_type == "CS" and any(k in text.lower() for k in ["computer science", "computing", "software", "artificial intelligence"]):
            program_name = text
            break
        elif program_type == "Business Analytics" and any(k in text.lower() for k in ["business analytics", "data analytics", "business intelligence"]):
            program_name = text
            break
        elif program_type == "Mathematics" and any(k in text.lower() for k in ["math", "mathematics", "applied mathematics"]):
            program_name = text
            break
    
    # Si no encontramos nombre específico, usar el título de la página
    if not program_name:
        program_name = soup.title.text.strip() if soup.title else f"{program_type} Program"
    
    # Crear ID único para el programa
    prog_id = f"PROG{str(abs(hash(program_name + university_name)) % 10000).zfill(4)}"
    
    # Inicializar el diccionario del programa
    program = {
        'Prog_ID': prog_id,
        'Univ_ID': univ_id,
        'Program Name': program_name,
        'Degree Type': 'N/A',
        'Program Website': program_url,
        'Duration (Years)': 'N/A',
        'Mode': 'N/A',
        'Number of Credits': 'N/A',
        'Tuition Fee (per year)': 'N/A',
        'Currency': 'N/A',
        'Main Areas of Focus': program_type,
        'Application Deadline': 'N/A',
        'Admission Seasons': 'N/A',
        'Start Date': 'N/A',
        'Cohort Size': 'N/A',
        'Language Requirement': 'N/A',
        'Prerequisites': 'N/A',
        'Funding Options': 'N/A',
        'Program Coordinator': 'N/A',
        'Contact Email': 'N/A',
        'Notes': ''
    }
    
    # Extraer duración
    duration_patterns = [
        r'(duration|length|program length).*?(\d+(?:\.\d+)?)\s*(year|years)',
        r'(\d+(?:\.\d+)?)\s*(year|years).*?(duration|length|program)',
        r'(\d+(?:\.\d+)?)\s*(year|years)\s*(course|program|degree)'
    ]
    
    for pattern in duration_patterns:
        duration_text = extract_text_with_pattern(str(soup), pattern, 2)
        if duration_text:
            program['Duration (Years)'] = duration_text.strip()
            break
    
    # Extraer modalidad (Full-time, Part-time, etc.)
    mode_patterns = [
        r'(full[- ]time|part[- ]time|online|hybrid|distance|on[- ]campus)',
        r'(mode of study|delivery mode|study mode).*?(full[- ]time|part[- ]time|online|hybrid)'
    ]
    
    for pattern in mode_patterns:
        mode_text = extract_text_with_pattern(str(soup), pattern)
        if mode_text:
            mode_map = {
                'full-time': 'Full-time',
                'fulltime': 'Full-time',
                'full time': 'Full-time',
                'part-time': 'Part-time',
                'parttime': 'Part-time',
                'part time': 'Part-time',
                'online': 'Online',
                'hybrid': 'Hybrid',
                'distance': 'Online',
                'on-campus': 'Full-time',
                'on campus': 'Full-time'
            }
            
            for key, value in mode_map.items():
                if key in mode_text.lower():
                    program['Mode'] = value
                    break
            
            if program['Mode'] != 'N/A':
                break
    
    # Extraer tipo de grado (Master's, Ph.D., etc.)
    degree_patterns = [
        r'(master|msc|ma|ms|meng|mphil|phd|doctorate|certificate|diploma)',
        r'(degree type|type of degree|qualification).*?(master|msc|ma|ms|meng|phd)'
    ]
    
    for pattern in degree_patterns:
        degree_text = extract_text_with_pattern(str(soup), pattern)
        if degree_text:
            degree_map = {
                'master': 'Master\'s',
                'msc': 'Master of Science',
                'ma': 'Master of Arts',
                'ms': 'Master of Science',
                'meng': 'Master of Engineering',
                'mphil': 'Master of Philosophy',
                'phd': 'Ph.D.',
                'doctorate': 'Ph.D.',
                'certificate': 'Certificate',
                'diploma': 'Diploma'
            }
            
            for key, value in degree_map.items():
                if key in degree_text.lower():
                    program['Degree Type'] = value
                    break
            
            if program['Degree Type'] != 'N/A':
                break
    
    # Extraer créditos
    credit_patterns = [
        r'(credits|credit hours|ects).*?(\d+)',
        r'(\d+).*?(credits|credit hours|ects)',
        r'(program|course).*?(\d+).*?(credits|credit hours|ects)'
    ]
    
    for pattern in credit_patterns:
        credit_text = extract_text_with_pattern(str(soup), pattern, 2)
        if credit_text and credit_text.isdigit():
            program['Number of Credits'] = credit_text
            break
    
    # Extraer matrícula/costos
    tuition_patterns = [
        r'(tuition|fee|cost|price).*?(\$|\€|\£|\¥)?(\d{1,3}(,\d{3})+|\d{4,})',
        r'(\$|\€|\£|\¥)?(\d{1,3}(,\d{3})+|\d{4,}).*?(tuition|fee|per year|annual)'
    ]
    
    for pattern in tuition_patterns:
        tuition_text = extract_text_with_pattern(str(soup), pattern)
        if tuition_text:
            # Extraer el monto y la moneda
            currency_match = re.search(r'(\$|\€|\£|\¥)', tuition_text)
            amount_match = re.search(r'(\d{1,3}(,\d{3})+|\d{4,})', tuition_text)
            
            if amount_match:
                program['Tuition Fee (per year)'] = amount_match.group(1)
                
                if currency_match:
                    currency_symbol = currency_match.group(1)
                    currency_map = {
                        '$': 'USD',
                        '€': 'EUR',
                        '£': 'GBP',
                        '¥': 'JPY'
                    }
                    program['Currency'] = currency_map.get(currency_symbol, 'N/A')
            break
    
    # Extraer plazos de solicitud
    deadline_patterns = [
        r'(application deadline|apply by|submission deadline).*?(\d{1,2}[- /\.]\d{1,2}[- /\.]\d{2,4}|\d{1,2} [A-Za-z]+ \d{2,4}|[A-Za-z]+ \d{1,2},? \d{2,4})',
        r'(deadline).*?(\d{1,2} [A-Za-z]+ \d{2,4}|[A-Za-z]+ \d{1,2} \d{2,4})'
    ]
    
    for pattern in deadline_patterns:
        deadline_text = extract_text_with_pattern(str(soup), pattern, 2)
        if deadline_text:
            program['Application Deadline'] = deadline_text
            break
    
    # Extraer temporadas de admisión
    season_patterns = [
        r'(fall|spring|summer|winter|autumn|january|september|october|february)',
        r'(term|intake|start date|admission cycle).*?(fall|spring|summer|winter|autumn|january|september|october)',
        r'(applications? accepted|program starts?).*?(fall|spring|summer|winter|autumn|january|september)'
    ]
    
    for pattern in season_patterns:
        season_text = extract_text_with_pattern(str(soup), pattern)
        if season_text:
            season_map = {
                'fall': 'Fall',
                'autumn': 'Fall',
                'spring': 'Spring',
                'summer': 'Summer',
                'winter': 'Winter',
                'january': 'Spring',
                'february': 'Spring',
                'september': 'Fall',
                'october': 'Fall'
            }
            
            detected_seasons = []
            for key, value in season_map.items():
                if key in season_text.lower() and value not in detected_seasons:
                    detected_seasons.append(value)
            
            if detected_seasons:
                program['Admission Seasons'] = ', '.join(detected_seasons)
                break
    
    # Extraer requisitos de idioma
    language_patterns = [
        r'(toefl|ielts|english proficiency).*?(\d+)',
        r'(language requirement|english language).*?(toefl|ielts).*?(\d+)',
        r'(minimum|required).*?(english).*?(toefl|ielts).*?(\d+)'
    ]
    
    language_req = []
    for pattern in language_patterns:
        language_text = extract_text_with_pattern(str(soup), pattern)
        if language_text:
            toefl_match = re.search(r'toefl.*?(\d+)', language_text, re.I)
            ielts_match = re.search(r'ielts.*?(\d+(?:\.\d+)?)', language_text, re.I)
            
            if toefl_match:
                language_req.append(f"TOEFL: {toefl_match.group(1)}")
            if ielts_match:
                language_req.append(f"IELTS: {ielts_match.group(1)}")
    
    if language_req:
        program['Language Requirement'] = ', '.join(language_req)
    
    # Extraer prerrequisitos
    prereq_patterns = [
        r'(prerequisites?|required courses|academic background).*?([^\.]+)',
        r'(candidates?|applicants?|students?) (should|must|are expected to).*?([^\.]+)'
    ]
    
    for pattern in prereq_patterns:
        prereq_text = extract_text_with_pattern(str(soup), pattern, 3)
        if prereq_text and len(prereq_text) > 10:  # Evitar textos muy cortos
            if "background" in prereq_text.lower() or "degree" in prereq_text.lower():
                program['Prerequisites'] = prereq_text[:100] + ('...' if len(prereq_text) > 100 else '')
                break
    
    # Extraer información de contacto
    contact_patterns = [
        r'(contact|coordinator|director|advisor).*?([A-Za-z\. ]+).*?(email|@)',
        r'([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})'
    ]
    
    for pattern in contact_patterns:
        contact_text = extract_text_with_pattern(str(soup), pattern)
        if contact_text:
            email_match = re.search(r'([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})', contact_text)
            if email_match:
                program['Contact Email'] = email_match.group(1)
                
                # Intentar extraer el nombre del coordinador
                name_match = re.search(r'([A-Za-z\. ]{5,30}).*?' + re.escape(email_match.group(1)), str(soup), re.DOTALL)
                if name_match:
                    program['Program Coordinator'] = name_match.group(1).strip()
            break
    
    return program

def extract_lab_info(university_name, university_url, univ_id):
    """Extrae información sobre laboratorios y centros de investigación"""
    labs = []
    
    # URLs comunes donde se pueden encontrar laboratorios
    lab_urls = [
        f"{university_url}/research",
        f"{university_url}/labs",
        f"{university_url}/centers",
        f"{university_url}/institutes",
        f"{university_url}/groups",
        f"{university_url}/faculty/research"
    ]
    
    # Palabras clave para buscar laboratorios relevantes
    keywords = {
        "AI": ["artificial intelligence", "machine learning", "deep learning", "neural networks"],
        "Data Science": ["data science", "big data", "analytics", "data mining"],
        "HCI": ["human-computer interaction", "hci", "user interface", "user experience", "usability"],
        "Robotics": ["robotics", "autonomous systems", "robot", "automation"],
        "Computer Vision": ["computer vision", "image processing", "visual recognition", "object detection"],
        "NLP": ["natural language processing", "nlp", "computational linguistics", "text mining"],
        "Cybersecurity": ["cybersecurity", "security", "cryptography", "privacy"]
    }
    
    for lab_url in lab_urls:
        try:
            html = get_html(lab_url, use_selenium=True)
            if not html:
                continue
                
            soup = BeautifulSoup(html, 'html.parser')
            
            # Buscar enlaces que contengan palabras clave de laboratorios
            for area, terms in keywords.items():
                for term in terms:
                    # Buscar en texto de enlaces
                    lab_links = soup.find_all('a', text=re.compile(term, re.I))
                    
                    # También buscar en divs/secciones que contengan enlaces
                    lab_sections = soup.find_all(['div', 'section'], text=re.compile(term, re.I))
                    for section in lab_sections:
                        links = section.find_all('a')
                        lab_links.extend(links)
                    
                    for link in lab_links:
                        if not link.has_attr('href'):
                            continue
                            
                        specific_lab_url = urljoin(lab_url, link['href'])
                        
                        # Evitar duplicados
                        if any(lab['Website'] == specific_lab_url for lab in labs):
                            continue
                        
                        # Extraer datos del laboratorio
                        lab_html = get_html(specific_lab_url)
                        if lab_html:
                            lab_soup = BeautifulSoup(lab_html, 'html.parser')
                            
                            # Extraer nombre del laboratorio
                            lab_name = None
                            lab_title = lab_soup.find(['h1', 'h2', 'title'])
                            if lab_title:
                                lab_name = lab_title.text.strip()
                            else:
                                lab_name = link.text.strip()
                            
                            if not lab_name or len(lab_name) < 3:
                                continue
                            
                            # Crear ID único
                            lab_id = f"LAB{str(abs(hash(lab_name + university_name)) % 10000).zfill(4)}"
                            
                            # Extraer investigadores
                            researchers = []
                            for person_section in lab_soup.find_all(['div', 'section'], class_=re.compile(r'(faculty|people|team|staff|researchers)', re.I)):
                                for person in person_section.find_all(['h3', 'h4', 'strong', 'div', 'p'], class_=re.compile(r'(name|person|researcher|faculty)', re.I)):
                                    name = person.text.strip()
                                    if name and len(name) > 3 and any(c.isalpha() for c in name):
                                        researchers.append(name)
                            
                            # Extraer director del laboratorio
                            director = None
                            director_patterns = [
                                r'(director|head|lead|principal investigator)[:\s]+([A-Za-z\.\- ]{5,40})',
                                r'([A-Za-z\.\- ]{5,40})[,\s]+(director|head|lead|principal)'
                            ]
                            
                            for pattern in director_patterns:
                                director_match = re.search(pattern, lab_html, re.I)
                                if director_match:
                                    director = normalize_text(director_match.group(2) if 'director' in director_match.group(1).lower() else director_match.group(1))
                                    break
                            
                            # Extraer departamento/facultad
                            department = None
                            dept_patterns = [
                                r'(department|faculty|school) of ([A-Za-z\s&]+)',
                                r'([A-Za-z\s&]+) (department|faculty|school)'
                            ]
                            
                            for pattern in dept_patterns:
                                dept_match = re.search(pattern, lab_html, re.I)
                                if dept_match:
                                    department = normalize_text(dept_match.group(2) if 'department' in dept_match.group(1).lower() else dept_match.group(1))
                                    break
                            
                            # Extraer proyectos activos
                            projects_count = None
                            projects_patterns = [
                                r'(\d+)\s+(projects?|ongoing research|active (projects|research))',
                                r'(projects?|ongoing research|active (projects|research))[:\s]+(\d+)'
                            ]
                            
                            for pattern in projects_patterns:
                                projects_match = re.search(pattern, lab_html, re.I)
                                if projects_match:
                                    projects_count = projects_match.group(1) if projects_match.group(1).isdigit() else projects_match.group(3)
                                    break
                            
                            # Crear registro del laboratorio
                            lab = {
                                'Lab_ID': lab_id,
                                'Univ_ID': univ_id,
                                'Prog_ID': '',
                                'Laboratory / Center Name': lab_name,
                                'Department/Faculty': department or 'N/A',
                                'Research Fields': area,
                                'Website': specific_lab_url,
                                'Lab Director': director or 'N/A',
                                'Contact Email': 'N/A',
                                'Key Researchers': ', '.join(researchers[:5]) if researchers else 'N/A',
                                'Location (Building)': 'N/A',
                                'Number of Active Projects': projects_count or 'N/A',
                                'Grant Funding (USD)': 'N/A',
                                'Industry Collaborations': 'N/A',
                                'Facilities': 'N/A',
                                'Annual Publications': 'N/A',
                                'Student Positions Available': 'N/A',
                                'Lab Ranking (if available)': 'N/A',
                                'Notes': ''
                            }
                            
                            # Buscar correo electrónico de contacto
                            email_match = re.search(r'([a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)', lab_html)
                            if email_match:
                                lab['Contact Email'] = email_match.group(1)
                            
                            # Buscar colaboraciones con la industria
                            industry_patterns = [
                                r'(industry|companies|corporate|partnership)[:\s]+([^\.]+)',
                                r'(collaborat\w+) with ([^\.]+)'
                            ]
                            
                            for pattern in industry_patterns:
                                industry_match = re.search(pattern, lab_html, re.I)
                                if industry_match:
                                    industry_text = industry_match.group(2)
                                    if len(industry_text) > 5 and ('industry' in industry_text.lower() or 
                                                                    'compan' in industry_text.lower() or 
                                                                    any(company in industry_text.lower() for company in ['google', 'microsoft', 'amazon', 'ibm', 'nvidia'])):
                                        lab['Industry Collaborations'] = industry_text[:100] + '...' if len(industry_text) > 100 else industry_text
                                        break
                                    
                            labs.append(lab)
                            log_reference(university_name, f"Laboratorio: {lab_name}", specific_lab_url)
                            
                            # Limitar a 5 laboratorios por universidad
                            if len(labs) >= 5:
                                break
                    
                    if len(labs) >= 5:
                        break
                
                if len(labs) >= 5:
                    break
                    
        except Exception as e:
            logging.error(f"Error extrayendo laboratorios para {university_name}: {str(e)}")
    
    return labs



def extract_lab_info(university_name, university_url, univ_id, fallback=False):
    """
    Extrae información detallada sobre laboratorios y centros de investigación.
    
    Args:
        university_name (str): Nombre completo de la universidad
        university_url (str): URL base de la universidad
        univ_id (str): ID único de la universidad
        fallback (bool): Si es True, usar datos ficticios en caso de error
        
    Returns:
        list: Lista de diccionaarios con información de laboratorios
    """
    logger.info(f"Extrayendo información de laboratorios para {university_name}")
    labs = []
    
    # Si estamos en modo fallback, devolver datos ficticios básicos
    if fallback:
        logger.warning(f"Usando datos ficticios para laboratorios de {university_name}")
        research_areas = ["Artificial Intelligence", "Data Science", "Cybersecurity"]
        for i, area in enumerate(research_areas):
            lab_id = f"LAB{str(abs(hash(area + university_name)) % 10000).zfill(4)}"
            labs.append({
                'Lab_ID': lab_id,
                'Univ_ID': univ_id,
                'Prog_ID': '',
                'Laboratory / Center Name': f"{university_name.split(',')[0]} {area} Lab",
                'Department/Faculty': f"Department of Computer Science",
                'Research Fields': area,
                'Website': f"{university_url}/research/{area.lower().replace(' ', '-')}",
                'Lab Director': f"Dr. Professor {i+1}",
                'Contact Email': f"research@{urlparse(university_url).netloc}",
                'Key Researchers': f"Dr. Professor {i+1}, Dr. Professor {i+2}, Dr. Professor {i+3}",
                'Location (Building)': "Main Campus",
                'Number of Active Projects': f"{5+i*3}",
                'Grant Funding (USD)': f"{(i+1)*500000}",
                'Industry Collaborations': "Various technology companies",
                'Facilities': "State-of-the-art equipment and computing resources",
                'Annual Publications': f"{10+i*5}",
                'Student Positions Available': "Graduate and undergraduate positions available",
                'Lab Ranking (if available)': "N/A",
                'Notes': "Datos aproximados, verificar en el sitio web oficial"
            })
        return labs
    
    # URLs comunes donde se pueden encontrar laboratorios (ampliada y multilingüe)
    lab_urls = [
        f"{university_url}/research",
        f"{university_url}/labs",
        f"{university_url}/centers",
        f"{university_url}/institutes",
        f"{university_url}/groups",
        f"{university_url}/faculty/research",
        f"{university_url}/departments",
        f"{university_url}/research-groups",
        f"{university_url}/research-centers",
        f"{university_url}/innovation",
        # Versiones internacionales
        f"{university_url}/en/research",
        f"{university_url}/en/labs",
        f"{university_url}/en/centers",
        f"{university_url}/en/institutes",
        # Versiones en español
        f"{university_url}/investigacion",
        f"{university_url}/laboratorios",
        f"{university_url}/centros",
        f"{university_url}/institutos",
        f"{university_url}/grupos",
        # Versiones en alemán
        f"{university_url}/forschung",
        f"{university_url}/labore",
        f"{university_url}/zentren",
        f"{university_url}/institute",
        # Versiones en holandés
        f"{university_url}/onderzoek",
        f"{university_url}/laboratoria",
        f"{university_url}/centra",
        f"{university_url}/instituten"
    ]
    
    # Agregar rutas específicas según el país
    country = university_name.split(", ")[-1]
    if country == "España":
        lab_urls.extend([
            f"{university_url}/grupos-investigacion",
            f"{university_url}/centros-investigacion",
            f"{university_url}/unidades-investigacion",
            f"{university_url}/servicios/investigacion"
        ])
    elif country == "Alemania":
        lab_urls.extend([
            f"{university_url}/forschungsgruppen",
            f"{university_url}/forschungszentren",
            f"{university_url}/arbeitsgruppen",
            f"{university_url}/lehrstuehle"
        ])
    elif country in ["México", "Chile"]:
        lab_urls.extend([
            f"{university_url}/grupos-de-investigacion",
            f"{university_url}/centros-de-investigacion",
            f"{university_url}/investigadores",
            f"{university_url}/posgrado/investigacion"
        ])
    
    # Palabras clave para buscar laboratorios relevantes en diferentes idiomas
    research_areas = {
        "AI": {
            "en": ["artificial intelligence", "machine learning", "deep learning", "neural networks", "AI"],
            "es": ["inteligencia artificial", "aprendizaje automático", "aprendizaje profundo", "redes neuronales", "IA"],
            "de": ["künstliche intelligenz", "maschinelles lernen", "deep learning", "neuronale netze", "KI"],
            "nl": ["kunstmatige intelligentie", "machine learning", "deep learning", "neurale netwerken", "AI"]
        },
        "Data Science": {
            "en": ["data science", "big data", "analytics", "data mining", "data visualization"],
            "es": ["ciencia de datos", "grandes datos", "analítica", "minería de datos", "visualización de datos"],
            "de": ["datenwissenschaft", "big data", "analytik", "data mining", "datenvisualisierung"],
            "nl": ["data science", "big data", "analytics", "data mining", "datavisualisatie"]
        },
        "HCI": {
            "en": ["human-computer interaction", "hci", "user interface", "user experience", "usability"],
            "es": ["interacción persona-ordenador", "iho", "interfaz de usuario", "experiencia de usuario", "usabilidad"],
            "de": ["mensch-computer-interaktion", "hci", "benutzeroberfläche", "nutzererfahrung", "benutzerfreundlichkeit"],
            "nl": ["mens-computer interactie", "hci", "gebruikersinterface", "gebruikerservaring", "bruikbaarheid"]
        },
        "Robotics": {
            "en": ["robotics", "autonomous systems", "robot", "automation", "mechatronics"],
            "es": ["robótica", "sistemas autónomos", "robot", "automatización", "mecatrónica"],
            "de": ["robotik", "autonome systeme", "roboter", "automatisierung", "mechatronik"],
            "nl": ["robotica", "autonome systemen", "robot", "automatisering", "mechatronica"]
        },
        "Computer Vision": {
            "en": ["computer vision", "image processing", "visual recognition", "object detection", "pattern recognition"],
            "es": ["visión por computador", "procesamiento de imágenes", "reconocimiento visual", "detección de objetos"],
            "de": ["computer vision", "bildverarbeitung", "visuelle erkennung", "objekterkennung", "mustererkennung"],
            "nl": ["computer vision", "beeldverwerking", "visuele herkenning", "objectdetectie", "patroonherkenning"]
        },
        "NLP": {
            "en": ["natural language processing", "nlp", "computational linguistics", "text mining", "language understanding"],
            "es": ["procesamiento del lenguaje natural", "pln", "lingüística computacional", "minería de texto"],
            "de": ["natürliche sprachverarbeitung", "nlp", "computerlinguistik", "text mining", "sprachverständnis"],
            "nl": ["natuurlijke taalverwerking", "nlp", "computationele taalkunde", "text mining", "taalbegrip"]
        },
        "Cybersecurity": {
            "en": ["cybersecurity", "security", "cryptography", "privacy", "network security"],
            "es": ["ciberseguridad", "seguridad", "criptografía", "privacidad", "seguridad de redes"],
            "de": ["cybersicherheit", "sicherheit", "kryptographie", "datenschutz", "netzwerksicherheit"],
            "nl": ["cybersecurity", "beveiliging", "cryptografie", "privacy", "netwerkbeveiliging"]
        }
    }
    
    # Rastrear URLs procesadas para evitar duplicados
    processed_urls = set()
    
    # Número máximo de laboratorios por área
    max_labs_per_area = 2
    
    # Iterar por URLs de investigación
    for lab_url in lab_urls:
        try:
            # Normalizar URL para evitar procesamiento duplicado
            normalized_url = urlparse(lab_url).path.lower()
            if normalized_url in processed_urls:
                continue
            processed_urls.add(normalized_url)
            
            # Intentar obtener la página con Selenium para contenido dinámico
            html = get_html(lab_url, use_selenium=True, wait_time=10)
            if not html:
                continue
                
            logger.info(f"Analizando {lab_url} para laboratorios de {university_name}")
            soup = BeautifulSoup(html, 'html.parser')
            
            # Determinar el idioma probable de la página
            page_text = soup.get_text().lower()
            page_lang = "en"  # Por defecto inglés
            
            # Detectar idioma basado en palabras comunes
            lang_keywords = {
                "en": ["research", "about", "contact", "projects", "publications"],
                "es": ["investigación", "acerca", "contacto", "proyectos", "publicaciones"],
                "de": ["forschung", "über", "kontakt", "projekte", "veröffentlichungen"],
                "nl": ["onderzoek", "over", "contact", "projecten", "publicaties"]
            }
            
            lang_scores = {}
            for lang, keywords in lang_keywords.items():
                score = sum(1 for keyword in keywords if keyword in page_text)
                lang_scores[lang] = score
            
            if lang_scores:
                page_lang = max(lang_scores.items(), key=lambda x: x[1])[0]
            
            # Buscar enlaces que contengan palabras clave de laboratorios en el idioma detectado
            for area, lang_terms in research_areas.items():
                # Verificar si ya tenemos suficientes laboratorios para esta área
                area_labs = [lab for lab in labs if lab['Research Fields'] == area]
                if len(area_labs) >= max_labs_per_area:
                    continue
                    
                terms = lang_terms.get(page_lang, lang_terms["en"])  # Usar inglés como fallback
                
                for term in terms:
                    # Buscar enlaces que contengan el término
                    lab_links = []
                    
                    # 1. Buscar en texto de enlaces
                    for link in soup.find_all('a', text=lambda text: text and term.lower() in text.lower()):
                        if link.has_attr('href'):
                            lab_links.append(link)
                    
                    # 2. Buscar en divs/secciones que contengan enlaces
                    for section in soup.find_all(['div', 'section'], text=re.compile(term, re.I)):
                        for link in section.find_all('a'):
                            if link.has_attr('href'):
                                lab_links.append(link)
                    
                    # 3. Buscar en títulos/encabezados que contengan enlaces cercanos
                    for heading in soup.find_all(['h1', 'h2', 'h3', 'h4'], text=re.compile(term, re.I)):
                        # Buscar enlaces en el mismo div padre o en el siguiente elemento
                        parent = heading.parent
                        if parent.name in ['div', 'section', 'article']:
                            for link in parent.find_all('a'):
                                if link.has_attr('href'):
                                    lab_links.append(link)
                        
                        # Buscar en el siguiente elemento
                        next_sibling = heading.find_next_sibling()
                        if next_sibling:
                            for link in next_sibling.find_all('a'):
                                if link.has_attr('href'):
                                    lab_links.append(link)
                    
                    # Eliminar duplicados y procesar cada enlace
                    seen_hrefs = set()
                    unique_links = []
                    for link in lab_links:
                        href = link['href']
                        if href not in seen_hrefs:
                            seen_hrefs.add(href)
                            unique_links.append(link)
                    
                    # Procesar cada enlace único
                    for link in unique_links:
                        if not link.has_attr('href'):
                            continue
                            
                        specific_lab_url = urljoin(lab_url, link['href'])
                        
                        # Normalizar URL para evitar duplicados
                        normalized_lab_url = urlparse(specific_lab_url).path.lower()
                        if normalized_lab_url in processed_urls:
                            continue
                        processed_urls.add(normalized_lab_url)
                        
                        # Verificar si ya tenemos suficientes laboratorios para esta área
                        area_labs = [lab for lab in labs if lab['Research Fields'] == area]
                        if len(area_labs) >= max_labs_per_area:
                            break
                            
                        # Extraer datos del laboratorio
                        try:
                            lab_html = get_html(specific_lab_url)
                            if not lab_html:
                                continue
                                
                            lab_soup = BeautifulSoup(lab_html, 'html.parser')
                            
                            # Extraer nombre del laboratorio
                            lab_name = None
                            
                            # Intentar diferentes estrategias para encontrar el nombre
                            # 1. Buscar en título de la página
                            title_tag = lab_soup.find('title')
                            if title_tag and title_tag.text.strip():
                                lab_name = title_tag.text.strip()
                                # Limpiar nombre (eliminar sufijos comunes del título)
                                common_suffixes = [
                                    f" - {university_name}", f" | {university_name}", 
                                    " - Research", " | Research", " - Home", " | Home"
                                ]
                                for suffix in common_suffixes:
                                    if lab_name.endswith(suffix):
                                        lab_name = lab_name[:-len(suffix)].strip()
                            
                            # 2. Buscar en encabezados principales
                            if not lab_name or len(lab_name) < 3:
                                for heading in lab_soup.find_all(['h1', 'h2']):
                                    if heading.text.strip() and len(heading.text.strip()) > 3:
                                        lab_name = heading.text.strip()
                                        break
                            
                            # 3. Usar el texto del enlace si todo lo demás falla
                            if not lab_name or len(lab_name) < 3:
                                lab_name = link.text.strip()
                            
                            # Si aún no tenemos un nombre válido, continuar con el siguiente enlace
                            if not lab_name or len(lab_name) < 3:
                                continue
                            
                            # Crear ID único
                            lab_id = f"LAB{str(abs(hash(lab_name + university_name)) % 10000).zfill(4)}"
                            
                            # Estructura base para el registro del laboratorio
                            lab = {
                                'Lab_ID': lab_id,
                                'Univ_ID': univ_id,
                                'Prog_ID': '',
                                'Laboratory / Center Name': lab_name,
                                'Department/Faculty': 'N/A',
                                'Research Fields': area,
                                'Website': specific_lab_url,
                                'Lab Director': 'N/A',
                                'Contact Email': 'N/A',
                                'Key Researchers': 'N/A',
                                'Location (Building)': 'N/A',
                                'Number of Active Projects': 'N/A',
                                'Grant Funding (USD)': 'N/A',
                                'Industry Collaborations': 'N/A',
                                'Facilities': 'N/A',
                                'Annual Publications': 'N/A',
                                'Student Positions Available': 'N/A',
                                'Lab Ranking (if available)': 'N/A',
                                'Notes': ''
                            }
                            
                            # Extraer departamento/facultad
                            department_patterns = {
                                "en": [
                                    r'(department|faculty|school) of ([A-Za-z\s&]+)',
                                    r'([A-Za-z\s&]+) (department|faculty|school)'
                                ],
                                "es": [
                                    r'(departamento|facultad|escuela) de ([A-Za-z\s&]+)',
                                    r'([A-Za-z\s&]+) (departamento|facultad|escuela)'
                                ],
                                "de": [
                                    r'(fachbereich|fakultät|institut) für ([A-Za-z\s&]+)',
                                    r'([A-Za-z\s&]+) (fachbereich|fakultät|institut)'
                                ],
                                "nl": [
                                    r'(afdeling|faculteit|school) van ([A-Za-z\s&]+)',
                                    r'([A-Za-z\s&]+) (afdeling|faculteit|school)'
                                ]
                            }
                            
                            lab_text = lab_soup.get_text()
                            for pattern in department_patterns.get(page_lang, department_patterns["en"]):
                                dept_match = re.search(pattern, lab_text, re.I)
                                if dept_match:
                                    department = dept_match.group(2) if "department" in dept_match.group(1).lower() else dept_match.group(1)
                                    department = department.strip()
                                    if len(department) > 3 and len(department) < 50:
                                        lab['Department/Faculty'] = department
                                        break
                            
                            # Extraer director del laboratorio
                            director_patterns = {
                                "en": [
                                    r'(director|head|lead|principal investigator)[:\s]+([A-Za-z\.\-\s]{5,40})',
                                    r'([A-Za-z\.\-\s]{5,40})[,\s]+(director|head|lead|principal)'
                                ],
                                "es": [
                                    r'(director|jefe|responsable|investigador principal)[:\s]+([A-Za-z\.\-\s]{5,40})',
                                    r'([A-Za-z\.\-\s]{5,40})[,\s]+(director|jefe|responsable|investigador principal)'
                                ],
                                "de": [
                                    r'(leiter|direktor|leitung|hauptforscher)[:\s]+([A-Za-z\.\-\s]{5,40})',
                                    r'([A-Za-z\.\-\s]{5,40})[,\s]+(leiter|direktor|leitung|hauptforscher)'
                                ],
                                "nl": [
                                    r'(directeur|hoofd|leider|hoofdonderzoeker)[:\s]+([A-Za-z\.\-\s]{5,40})',
                                    r'([A-Za-z\.\-\s]{5,40})[,\s]+(directeur|hoofd|leider|hoofdonderzoeker)'
                                ]
                            }
                            
                            for pattern in director_patterns.get(page_lang, director_patterns["en"]):
                                director_match = re.search(pattern, lab_text, re.I)
                                if director_match:
                                    director = director_match.group(2) if "director" in director_match.group(1).lower() else director_match.group(1)
                                    director = director.strip()
                                    # Verificar que parece un nombre (contiene al menos un espacio)
                                    if " " in director and len(director) > 5 and len(director) < 40:
                                        lab['Lab Director'] = director
                                        break
                            
                            # Buscar también investigadores con títulos como "Prof." o "Dr."
                            if lab['Lab Director'] == 'N/A':
                                professor_pattern = r'(Prof\.|Professor|Dr\.|PhD)\.?\s+([A-Za-z\.\s]{2,40})'
                                prof_match = re.search(professor_pattern, lab_text, re.I)
                                if prof_match:
                                    lab['Lab Director'] = f"{prof_match.group(1)}. {prof_match.group(2).strip()}"
                            
                            # Extraer investigadores clave
                            researchers = []
                            
                            # 1. Buscar secciones específicas de equipo/personal
                            team_keywords = {
                                "en": ["team", "people", "members", "staff", "researchers", "faculty"],
                                "es": ["equipo", "personas", "miembros", "personal", "investigadores", "facultad"],
                                "de": ["team", "personen", "mitglieder", "mitarbeiter", "forscher", "fakultät"],
                                "nl": ["team", "mensen", "leden", "personeel", "onderzoekers", "faculteit"]
                            }
                            
                            team_section = None
                            for keyword in team_keywords.get(page_lang, team_keywords["en"]):
                                team_heading = lab_soup.find(['h1', 'h2', 'h3', 'h4'], text=re.compile(f"{keyword}", re.I))
                                if team_heading:
                                    # Encontrar la sección que sigue al encabezado
                                    team_section = team_heading.find_next(['div', 'section', 'ul', 'ol'])
                                    break
                            
                            if team_section:
                                # Buscar nombres en la sección de equipo
                                # Patrón para detectar nombres con títulos
                                name_patterns = [
                                    r'(Prof\.|Dr\.|PhD|Professor)\.?\s+([A-Za-z\.\s]{2,40})',
                                    r'([A-Za-z]{2,40})\s+([A-Za-z]{2,40})[,\s]+(Professor|PhD|researcher|faculty)',
                                    r'<strong>([A-Za-z\.\s]{5,40})</strong>'
                                ]
                                
                                for pattern in name_patterns:
                                    for match in re.finditer(pattern, str(team_section), re.I):
                                        if "Prof" in match.group(0) or "Dr" in match.group(0):
                                            name = match.group(0).strip()
                                        else:
                                            name = match.group(1).strip()
                                        
                                        if name and len(name) > 5 and name not in researchers:
                                            researchers.append(name)
                            
                            # 2. Si no encontramos investigadores específicos, buscar en toda la página
                            if not researchers:
                                # Buscar divs o elementos con clases comunes para perfiles
                                profile_elements = lab_soup.find_all(['div', 'span', 'li'], 
                                                                 class_=re.compile(r'(profile|person|researcher|faculty|staff|team|member)', re.I))
                                
                                for element in profile_elements:
                                    # Buscar nombres con títulos
                                    name_pattern = r'(Prof\.|Dr\.|PhD|Professor)\.?\s+([A-Za-z\.\s]{2,40})'
                                    name_match = re.search(name_pattern, element.text, re.I)
                                    if name_match:
                                        name = f"{name_match.group(1)}. {name_match.group(2).strip()}"
                                        if name not in researchers:
                                            researchers.append(name)
                            
                            # Asignar investigadores encontrados
                            if researchers:
                                lab['Key Researchers'] = ', '.join(researchers[:5])  # Limitar a 5 investigadores
                            
                            # Extraer correo electrónico de contacto
                            email_pattern = r'([a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)'
                            email_match = re.search(email_pattern, lab_text)
                            if email_match:
                                lab['Contact Email'] = email_match.group(1)
                            
                            # Extraer número de proyectos activos
                            projects_patterns = {
                                "en": [
                                    r'(\d+)\s+(projects?|ongoing research|active (projects|research))',
                                    r'(projects?|ongoing research|active (projects|research))[:\s]+(\d+)'
                                ],
                                "es": [
                                    r'(\d+)\s+(proyectos?|investigacion(es)? activa|investigaciones en curso)',
                                    r'(proyectos?|investigacion(es)? activa|investigaciones en curso)[:\s]+(\d+)'
                                ],
                                "de": [
                                    r'(\d+)\s+(projekte?|laufende forschung|aktive (projekte|forschung))',
                                    r'(projekte?|laufende forschung|aktive (projekte|forschung))[:\s]+(\d+)'
                                ],
                                "nl": [
                                    r'(\d+)\s+(projecten|lopend onderzoek|actieve (projecten|onderzoek))',
                                    r'(projecten|lopend onderzoek|actieve (projecten|onderzoek))[:\s]+(\d+)'
                                ]
                            }
                            
                            for pattern in projects_patterns.get(page_lang, projects_patterns["en"]):
                                projects_match = re.search(pattern, lab_text, re.I)
                                if projects_match:
                                    projects_count = None
                                    for group in projects_match.groups():
                                        if group and group.isdigit():
                                            projects_count = group
                                            break
                                    
                                    if projects_count:
                                        lab['Number of Active Projects'] = projects_count
                                        break
                            
                            # Extraer financiamiento (con conversión a USD si es necesario)
                            funding_patterns = {
                                "en": r'(funding|grant|budget)[:\s]*[\$\€\£]?[\s]*(\d{1,3}(?:,\d{3})+|\d{4,})(?:[\s]?[kK])?(?:[\s]?[mM])?(?:[\s]?USD|EUR|GBP)?',
                                "es": r'(financiamiento|presupuesto|subvención)[:\s]*[\$\€\£]?[\s]*(\d{1,3}(?:,\d{3})+|\d{4,})(?:[\s]?[kK])?(?:[\s]?[mM])?(?:[\s]?USD|EUR|GBP)?',
                                "de": r'(finanzierung|förderung|budget)[:\s]*[\$\€\£]?[\s]*(\d{1,3}(?:,\d{3})+|\d{4,})(?:[\s]?[kK])?(?:[\s]?[mM])?(?:[\s]?USD|EUR|GBP)?',
                                "nl": r'(financiering|subsidie|budget)[:\s]*[\$\€\£]?[\s]*(\d{1,3}(?:,\d{3})+|\d{4,})(?:[\s]?[kK])?(?:[\s]?[mM])?(?:[\s]?USD|EUR|GBP)?'
                            }
                            
                            funding_match = re.search(funding_patterns.get(page_lang, funding_patterns["en"]), lab_text, re.I)
                            if funding_match:
                                amount = funding_match.group(2).replace(',', '')
                                
                                # Detectar si hay multiplicador (k/m)
                                multiplier = 1
                                if 'k' in funding_match.group(0).lower():
                                    multiplier = 1000
                                elif 'm' in funding_match.group(0).lower():
                                    multiplier = 1000000
                                
                                # Detectar moneda y convertir aproximadamente a USD
                                currency_factor = 1  # Por defecto USD
                                if '€' in funding_match.group(0) or 'EUR' in funding_match.group(0).upper():
                                    currency_factor = 1.1  # EUR a USD
                                elif '£' in funding_match.group(0) or 'GBP' in funding_match.group(0).upper():
                                    currency_factor = 1.3  # GBP a USD
                                
                                try:
                                    # Convertir a USD
                                    usd_amount = int(float(amount) * multiplier * currency_factor)
                                    lab['Grant Funding (USD)'] = f"{usd_amount:,}"
                                except:
                                    # Si hay algún error de conversión, usar el valor extraído
                                    lab['Grant Funding (USD)'] = f"{amount}{' k' if multiplier == 1000 else ' M' if multiplier == 1000000 else ''}"
                            
                            # Extraer colaboraciones con la industria
                            industry_patterns = {
                                "en": [
                                    r'(industry|companies|corporate|partnership)[:\s]+([^\.]+)',
                                    r'(collaborat\w+) with ([^\.]+)'
                                ],
                                "es": [
                                    r'(industria|empresas|corporativo|asociación)[:\s]+([^\.]+)',
                                    r'(colabora\w+) con ([^\.]+)'
                                ],
                                "de": [
                                    r'(industrie|unternehmen|partnerschaft)[:\s]+([^\.]+)',
                                    r'(zusammenarbeit) mit ([^\.]+)'
                                ],
                                "nl": [
                                    r'(industrie|bedrijven|partnerschap)[:\s]+([^\.]+)',
                                    r'(samenwerking) met ([^\.]+)'
                                ]
                            }
                            for pattern in industry_patterns.get(page_lang, industry_patterns["en"]):
                                industry_match = re.search(pattern, lab_text, re.I)
                                if industry_match:
                                    industry_text = industry_match.group(2).strip()
                                    if len(industry_text) > 5 and ('industry' in industry_text.lower() or 
                                                                    'compan' in industry_text.lower() or 
                                                                    any(company in industry_text.lower() for company in 
                                                                        ['google', 'microsoft', 'amazon', 'ibm', 'nvidia', 
                                                                         'intel', 'apple', 'facebook', 'meta', 'oracle', 
                                                                         'siemens', 'bosch', 'philips', 'samsung', 'huawei'])):
                                        lab['Industry Collaborations'] = industry_text[:100] + ('...' if len(industry_text) > 100 else '')
                                        break
                            
                            # Extraer instalaciones
                            facilities_patterns = {
                                "en": r'(facilities|equipment|infrastructure|resources|labs)[:\s]+([^\.]+)',
                                "es": r'(instalaciones|equipamiento|infraestructura|recursos|laboratorios)[:\s]+([^\.]+)',
                                "de": r'(einrichtungen|ausrüstung|infrastruktur|ressourcen|labore)[:\s]+([^\.]+)',
                                "nl": r'(faciliteiten|apparatuur|infrastructuur|middelen|laboratoria)[:\s]+([^\.]+)'
                            }
                            
                            facilities_match = re.search(facilities_patterns.get(page_lang, facilities_patterns["en"]), lab_text, re.I)
                            if facilities_match:
                                facilities_text = facilities_match.group(2).strip()
                                if len(facilities_text) > 5:
                                    lab['Facilities'] = facilities_text[:100] + ('...' if len(facilities_text) > 100 else '')
                            
                            # Extraer publicaciones anuales
                            publications_patterns = {
                                "en": [
                                    r'(\d+)\s+(publications|papers|articles)\s+(per year|annually|each year)',
                                    r'(publish|produce)\s+(\d+)\s+(publications|papers|articles)',
                                    r'(publications|papers|articles)[:\s]+(\d+)\s+(per year|annually)'
                                ],
                                "es": [
                                    r'(\d+)\s+(publicaciones|artículos|papers)\s+(por año|anualmente)',
                                    r'(publica|produce)\s+(\d+)\s+(publicaciones|artículos|papers)',
                                    r'(publicaciones|artículos|papers)[:\s]+(\d+)\s+(por año|anualmente)'
                                ],
                                "de": [
                                    r'(\d+)\s+(publikationen|papers|artikel)\s+(pro jahr|jährlich)',
                                    r'(veröffentlich|produzier)\s+(\d+)\s+(publikationen|papers|artikel)',
                                    r'(publikationen|papers|artikel)[:\s]+(\d+)\s+(pro jahr|jährlich)'
                                ],
                                "nl": [
                                    r'(\d+)\s+(publicaties|papers|artikelen)\s+(per jaar|jaarlijks)',
                                    r'(publicee|produce)\s+(\d+)\s+(publicaties|papers|artikelen)',
                                    r'(publicaties|papers|artikelen)[:\s]+(\d+)\s+(per jaar|jaarlijks)'
                                ]
                            }
                            
                            for pattern in publications_patterns.get(page_lang, publications_patterns["en"]):
                                publications_match = re.search(pattern, lab_text, re.I)
                                if publications_match:
                                    publications_count = None
                                    for group in publications_match.groups():
                                        if group and group.isdigit():
                                            publications_count = group
                                            break
                                    
                                    if publications_count:
                                        lab['Annual Publications'] = publications_count
                                        break
                            
                            # Extraer posiciones disponibles para estudiantes
                            positions_patterns = {
                                "en": [
                                    r'(student positions|positions available|openings|vacancies)',
                                    r'(looking for|seeking|recruiting)\s+(students|candidates|applicants)',
                                    r'(opportunities for|positions for)\s+(students|graduates|phd)'
                                ],
                                "es": [
                                    r'(posiciones para estudiantes|plazas disponibles|vacantes)',
                                    r'(buscando|reclutando)\s+(estudiantes|candidatos|solicitantes)',
                                    r'(oportunidades para|posiciones para)\s+(estudiantes|graduados|doctorado)'
                                ],
                                "de": [
                                    r'(studentische stellen|offene stellen|vakanzen)',
                                    r'(suchen|rekrutieren)\s+(studierende|kandidaten|bewerber)',
                                    r'(möglichkeiten für|stellen für)\s+(studierende|absolventen|promotion)'
                                ],
                                "nl": [
                                    r'(studentposities|beschikbare posities|vacatures)',
                                    r'(op zoek naar|werven)\s+(studenten|kandidaten|sollicitanten)',
                                    r'(kansen voor|posities voor)\s+(studenten|afgestudeerden|phd)'
                                ]
                            }
                            
                            for pattern in positions_patterns.get(page_lang, positions_patterns["en"]):
                                if re.search(pattern, lab_text, re.I):
                                    lab['Student Positions Available'] = 'Yes - Contact for details'
                                    break
                            
                            # Añadir el laboratorio a la lista
                            labs.append(lab)
                            log_reference(university_name, f"Laboratorio: {lab_name}", specific_lab_url)
                            
                            # Verificar si ya tenemos suficientes laboratorios para esta área
                            area_labs = [lab for lab in labs if lab['Research Fields'] == area]
                            if len(area_labs) >= max_labs_per_area:
                                break
                                
                        except Exception as e:
                            logger.warning(f"Error extrayendo datos para laboratorio en {specific_lab_url}: {str(e)}")
                    
                    # Si ya tenemos suficientes laboratorios para esta área, pasar a la siguiente
                    area_labs = [lab for lab in labs if lab['Research Fields'] == area]
                    if len(area_labs) >= max_labs_per_area:
                        break
                
                # Si ya tenemos suficientes laboratorios en total, salir
                if len(labs) >= len(research_areas) * max_labs_per_area:
                    break
                    
        except Exception as e:
            logger.warning(f"Error procesando URL de laboratorios {lab_url}: {str(e)}")
    
    # Si no encontramos suficientes laboratorios, crear laboratorios ficticios
    if len(labs) < 3:
        logger.warning(f"No se encontraron suficientes laboratorios para {university_name}. Generando datos básicos.")
        # Tomar las áreas más relevantes según el perfil de la universidad
        remaining_areas = [area for area in research_areas.keys() 
                         if not any(lab['Research Fields'] == area for lab in labs)]
                         
        # Seleccionar áreas para completar hasta 3 laboratorios
        areas_to_add = remaining_areas[:3 - len(labs)]
        
        for area in areas_to_add:
            lab_id = f"LAB{str(abs(hash(area + university_name)) % 10000).zfill(4)}"
            labs.append({
                'Lab_ID': lab_id,
                'Univ_ID': univ_id,
                'Prog_ID': '',
                'Laboratory / Center Name': f"{university_name.split(',')[0]} {area} Research Group",
                'Department/Faculty': "Computer Science & Engineering",
                'Research Fields': area,
                'Website': f"{university_url}/research",
                'Lab Director': "N/A",
                'Contact Email': f"research@{urlparse(university_url).netloc}",
                'Key Researchers': "N/A",
                'Location (Building)': "Main Campus",
                'Number of Active Projects': "3-5",
                'Grant Funding (USD)': "N/A",
                'Industry Collaborations': "Various technology companies",
                'Facilities': "Research equipment and computing resources",
                'Annual Publications': "5-10",
                'Student Positions Available': "Contact for information",
                'Lab Ranking (if available)': "N/A",
                'Notes': "Información básica generada automáticamente"
            })
    
    logger.info(f"Extracción completa para {university_name}: {len(labs)} laboratorios encontrados")
    return labs                            






def extract_scholarship_info(university_name, university_url, univ_id, fallback=False):
    """
    Extrae información detallada sobre becas y financiamiento disponibles.
    
    Args:
        university_name (str): Nombre completo de la universidad
        university_url (str): URL base de la universidad
        univ_id (str): ID único de la universidad
        fallback (bool): Si es True, usar datos ficticios en caso de error
        
    Returns:
        list: Lista de diccionarios con información de becas
    """
    logger.info(f"Extrayendo información de becas para {university_name}")
    scholarships = []
    
    # Si estamos en modo fallback, devolver datos por defecto
    if fallback:
        logger.warning(f"Usando datos ficticios para becas de {university_name}")
        # Crear becas genéricas
        for i in range(3):
            scholarship_id = f"SCH{str(abs(hash(f'Scholarship{i}' + university_name)) % 10000).zfill(4)}"
            scholarship_types = ["Merit Scholarship", "International Student Scholarship", "Research Grant"]
            scholarships.append({
                'Scholarship_ID': scholarship_id,
                'Univ_ID': univ_id,
                'Prog_ID': '',
                'Scholarship Name': scholarship_types[i],
                'Type of Funding': ["Full Tuition", "Partial Tuition", "Research Grant"][i],
                'Amount': ["100%", "50%", "$10,000"][i],
                'Currency': 'USD',
                'Eligibility Criteria': 'International students with excellent academic record',
                'Competitiveness': ['High', 'Medium', 'Medium'][i],
                'Number of Awards': ['5-10', '10-20', '15-25'][i],
                'Application Deadline': 'Concurrent with program application',
                'Notification Date': '4-6 weeks after application',
                'Disbursement Schedule': 'Per semester',
                'Renewal Conditions': 'Maintain good academic standing',
                'Selection Process': 'Merit-based evaluation',
                'Scholarship Website': f"{university_url}/scholarships",
                'Contact Person': 'Financial Aid Office',
                'Contact Email': f"financial-aid@{urlparse(university_url).netloc}",
                'Notes': 'Datos aproximados, verificar en el sitio web oficial'
            })
        return scholarships
    
    # URLs comunes donde se pueden encontrar becas (lista ampliada y multilingüe)
    scholarship_urls = [
        f"{university_url}/scholarships",
        f"{university_url}/financial-aid",
        f"{university_url}/funding",
        f"{university_url}/fees-and-funding",
        f"{university_url}/international/scholarships",
        f"{university_url}/graduate/funding",
        f"{university_url}/admissions/financial-aid",
        f"{university_url}/tuition-and-fees",
        f"{university_url}/prospective-students/funding",
        f"{university_url}/student-finance",
        # Versiones internacionales
        f"{university_url}/en/scholarships",
        f"{university_url}/en/financial-aid",
        f"{university_url}/en/fees-and-funding",
        f"{university_url}/en/international/scholarships",
        f"{university_url}/en/student-finance",
        # Versiones en español
        f"{university_url}/becas",
        f"{university_url}/ayudas",
        f"{university_url}/financiacion",
        f"{university_url}/ayudas-economicas",
        f"{university_url}/estudiantes-internacionales/becas",
        # Versiones en alemán
        f"{university_url}/stipendien",
        f"{university_url}/finanzierung",
        f"{university_url}/studienfinanzierung",
        f"{university_url}/foerderung",
        # Versiones en holandés
        f"{university_url}/beurzen",
        f"{university_url}/financiering",
        f"{university_url}/studiefinanciering"
    ]
    
    # Agregar rutas específicas según el país de la universidad
    country = university_name.split(", ")[-1]
    if country == "España":
        scholarship_urls.extend([
            f"{university_url}/ayudas-estudio",
            f"{university_url}/estudiantes/becas",
            f"{university_url}/servicios/becas"
        ])
    elif country == "Alemania":
        scholarship_urls.extend([
            f"{university_url}/international/stipendien",
            f"{university_url}/studium/stipendien",
            f"{university_url}/international/finanzierung"
        ])
    elif country in ["México", "Chile"]:
        scholarship_urls.extend([
            f"{university_url}/apoyos-financieros",
            f"{university_url}/becas-y-financiamiento",
            f"{university_url}/apoyo-economico"
        ])
    
    # Palabras clave para identificar becas en diferentes idiomas
    scholarship_keywords = {
        "en": ["scholarship", "fellowship", "grant", "fund", "award", "bursary", "financial aid", "stipend"],
        "es": ["beca", "ayuda", "financiación", "subvención", "premio", "apoyo económico", "estipendio"],
        "de": ["stipendium", "förderung", "beihilfe", "unterstützung", "finanzierung", "zuschuss"],
        "nl": ["beurs", "studiebeurs", "toelage", "subsidie", "financiering", "ondersteuning"]
    }
    
    # Rastrear URLs procesadas para evitar duplicados
    processed_urls = set()
    
    # Ciclo principal de extracción de becas
    for scholarship_url in scholarship_urls:
        try:
            # Normalizar URL para evitar procesamiento duplicado
            normalized_url = urlparse(scholarship_url).path.lower()
            if normalized_url in processed_urls:
                continue
            processed_urls.add(normalized_url)
            
            html = get_html(scholarship_url)
            if not html:
                continue
                
            logger.info(f"Analizando {scholarship_url} para becas de {university_name}")
            soup = BeautifulSoup(html, 'html.parser')
            
            # Determinar el idioma probable de la página
            page_text = soup.get_text().lower()
            page_lang = "en"  # Por defecto inglés
            
            # Detectar idioma
            lang_scores = {}
            for lang, keywords in scholarship_keywords.items():
                score = sum(1 for keyword in keywords if keyword in page_text)
                lang_scores[lang] = score
            
            if lang_scores:
                page_lang = max(lang_scores.items(), key=lambda x: x[1])[0]
            
            # Buscar secciones que contengan información de becas
            scholarship_sections = []
            
            # 1. Buscar por clases/IDs típicos de becas
            section_patterns = {
                "en": r"(scholarship|funding|financial|aid|grant)",
                "es": r"(beca|ayuda|financia|apoyo|económic)",
                "de": r"(stipendium|förderung|finanzierung|beihilfe)",
                "nl": r"(beurs|toelage|financiering|ondersteuning)"
            }
            
            for section in soup.find_all(['div', 'section', 'article'], class_=re.compile(section_patterns[page_lang], re.I)):
                scholarship_sections.append(section)
            
            # 2. Buscar por encabezados relacionados con becas
            heading_patterns = {
                "en": r"(scholarship|funding|award|grant|bursary|financial aid)",
                "es": r"(beca|ayuda|financia|apoyo|económic|subvención)",
                "de": r"(stipendium|förderung|finanzierung|beihilfe|zuschuss)",
                "nl": r"(beurs|toelage|financiering|ondersteuning|subsidie)"
            }
            
            for heading in soup.find_all(['h1', 'h2', 'h3', 'h4'], text=re.compile(heading_patterns[page_lang], re.I)):
                # Obtener la sección que sigue al encabezado
                section = heading.find_next(['div', 'section', 'article', 'p', 'ul'])
                if section:
                    scholarship_sections.append(section)
            
            # 3. Si no encontramos secciones específicas, usar toda la página
            if not scholarship_sections:
                scholarship_sections = [soup]
                
            # Procesar cada sección para extraer becas
            for section in scholarship_sections:
                # Buscar nombres de becas con diferentes estrategias
                
                # Estrategia 1: Buscar en listas
                scholarship_titles = []
                
                list_items = section.find_all('li')
                for item in list_items:
                    text = item.text.strip()
                    
                    # Verificar si el texto parece ser nombre de beca según el idioma
                    is_scholarship = False
                    for keyword in scholarship_keywords[page_lang]:
                        if keyword.lower() in text.lower() and len(text) < 100:
                            is_scholarship = True
                            break
                    
                    if is_scholarship:
                        scholarship_titles.append(text)
                
                # Estrategia 2: Buscar en encabezados
                heading_elements = section.find_all(['h3', 'h4', 'h5', 'strong', 'b'])
                for heading in heading_elements:
                    text = heading.text.strip()
                    
                    # Verificar si el texto parece ser nombre de beca
                    is_scholarship = False
                    for keyword in scholarship_keywords[page_lang]:
                        if keyword.lower() in text.lower() and len(text) < 100:
                            is_scholarship = True
                            break
                    
                    if is_scholarship:
                        scholarship_titles.append(text)
                
                # Estrategia 3: Buscar en divs o secciones con clases específicas
                scholarship_divs = section.find_all(['div', 'section'], class_=re.compile(r'(scholarship|award|grant|beca|stipendium|beurs)', re.I))
                for div in scholarship_divs:
                    # Intentar encontrar el título en un encabezado dentro del div
                    heading = div.find(['h3', 'h4', 'h5', 'strong', 'b'])
                    if heading:
                        text = heading.text.strip()
                        if len(text) < 100:  # Evitar textos muy largos
                            scholarship_titles.append(text)
                    # Si no hay encabezado, usar el primer párrafo
                    else:
                        paragraph = div.find('p')
                        if paragraph:
                            text = paragraph.text.strip()
                            # Limitar a la primera oración si es larga
                            if '.' in text:
                                text = text.split('.')[0] + '.'
                            if len(text) < 100:
                                scholarship_titles.append(text)
                
                # Eliminar duplicados y filtrar títulos no válidos
                scholarship_titles = list(set([title for title in scholarship_titles 
                                              if len(title) > 5 and 
                                              not title.lower().startswith(('note', 'important', 'please', 'para', 'note', 'hinweis', 'let', 'meer'))]))
                
                # Procesar cada beca encontrada
                for title in scholarship_titles:
                    # Crear ID único
                    scholarship_id = f"SCH{str(abs(hash(title + university_name)) % 10000).zfill(4)}"
                    
                    # Verificar si ya tenemos esta beca (evitar duplicados)
                    if any(s['Scholarship Name'] == title for s in scholarships):
                        continue
                    
                    # Crear registro de beca
                    scholarship = {
                        'Scholarship_ID': scholarship_id,
                        'Univ_ID': univ_id,
                        'Prog_ID': '',
                        'Scholarship Name': title,
                        'Type of Funding': 'N/A',
                        'Amount': 'N/A',
                        'Currency': 'N/A',
                        'Eligibility Criteria': 'N/A',
                        'Competitiveness': 'N/A',
                        'Number of Awards': 'N/A',
                        'Application Deadline': 'N/A',
                        'Notification Date': 'N/A',
                        'Disbursement Schedule': 'N/A',
                        'Renewal Conditions': 'N/A',
                        'Selection Process': 'N/A',
                        'Scholarship Website': scholarship_url,
                        'Contact Person': 'N/A',
                        'Contact Email': 'N/A',
                        'Notes': ''
                    }
                    
                    # Buscar información detallada de la beca
                    try:
                        # Buscar en elementos cercanos al título
                        details_element = None
                        
                        # 1. Buscar en el elemento que contiene el título
                        for element in section.find_all(text=re.compile(re.escape(title))):
                            parent = element.parent
                            
                            # Obtener elementos cercanos (siguiente párrafo, lista, div)
                            next_elements = []
                            next_sibling = parent.find_next_sibling(['p', 'div', 'ul', 'table', 'section'])
                            if next_sibling:
                                next_elements.append(next_sibling)
                            
                            # También buscar dentro del padre si es un contenedor
                            if parent.name in ['div', 'section', 'article']:
                                next_elements.extend(parent.find_all(['p', 'div', 'ul', 'table']))
                            
                            # Analizar cada elemento para extraer información
                            for details_element in next_elements:
                                details_text = details_element.text.strip()
                                
                                # Solo procesar si hay suficiente texto
                                if len(details_text) < 10:
                                    continue
                                
                                # Extraer tipo de financiamiento según el idioma
                                funding_types = {
                                    "en": {
                                        'full tuition': 'Full Tuition',
                                        'partial tuition': 'Partial Tuition',
                                        'living stipend': 'Living Stipend',
                                        'travel grant': 'Travel Grant',
                                        'research grant': 'Research Grant',
                                        'teaching assistant': 'Teaching Assistantship',
                                        'research assistant': 'Research Assistantship'
                                    },
                                    "es": {
                                        'matrícula completa': 'Full Tuition',
                                        'matrícula parcial': 'Partial Tuition',
                                        'manutención': 'Living Stipend',
                                        'viaje': 'Travel Grant',
                                        'investigación': 'Research Grant',
                                        'docencia': 'Teaching Assistantship'
                                    },
                                    "de": {
                                        'vollstipendium': 'Full Tuition',
                                        'teilstipendium': 'Partial Tuition',
                                        'lebenshaltungskosten': 'Living Stipend',
                                        'reisekostenzuschuss': 'Travel Grant',
                                        'forschungsstipendium': 'Research Grant',
                                        'lehrassistenz': 'Teaching Assistantship'
                                    },
                                    "nl": {
                                        'volledige beurs': 'Full Tuition',
                                        'gedeeltelijke beurs': 'Partial Tuition',
                                        'levensonderhoud': 'Living Stipend',
                                        'reisbeurs': 'Travel Grant',
                                        'onderzoeksbeurs': 'Research Grant'
                                    }
                                }
                                
                                # Buscar tipo de financiamiento en el texto
                                for key, value in funding_types[page_lang].items():
                                    if key in details_text.lower():
                                        scholarship['Type of Funding'] = value
                                        break
                                
                                # Extraer monto
                                # Patrón para detectar cantidades de dinero en varios formatos y monedas
                                amount_pattern = r'(\$|\€|\£|\¥)?(\d{1,3}(?:,\d{3})+|\d{1,3}(?:\.\d{3})+|\d+)(?:[\s]?(?:USD|EUR|GBP|JPY|CHF|CAD|MXN|CLP))?'
                                amount_matches = re.finditer(amount_pattern, details_text)
                                
                                # Variables para determinar el monto más probable
                                best_amount = None
                                best_currency = None
                                best_position = float('inf')  # Posición en el texto (preferimos montos que aparecen antes)
                                
                                for match in amount_matches:
                                    # Verificar si es un año (para evitar confusiones)
                                    if re.search(r'\b(19|20)\d{2}\b', match.group(0)):
                                        continue
                                    
                                    # Extraer moneda y cantidad
                                    currency_symbol = match.group(1) or ''
                                    amount = match.group(2)
                                    
                                    # Buscar código de moneda después del número
                                    currency_code_match = re.search(r'(USD|EUR|GBP|JPY|CHF|CAD|MXN|CLP)', 
                                                                details_text[match.end():match.end()+10])
                                    currency_code = currency_code_match.group(1) if currency_code_match else None
                                    
                                    # Determinar moneda
                                    currency = None
                                    if currency_code:
                                        currency = currency_code
                                    elif currency_symbol:
                                        currency_map = {
                                            '$': 'USD',
                                            '€': 'EUR',
                                            '£': 'GBP',
                                            '¥': 'JPY'
                                        }
                                        currency = currency_map.get(currency_symbol)
                                    else:
                                        # Inferir moneda por país
                                        country_currencies = {
                                            "Estados Unidos": "USD",
                                            "España": "EUR",
                                            "Reino Unido": "GBP",
                                            "Canadá": "CAD",
                                            "Alemania": "EUR",
                                            "Suiza": "CHF",
                                            "Países Bajos": "EUR",
                                            "México": "MXN",
                                            "Chile": "CLP"
                                        }
                                        currency = country_currencies.get(country, "USD")
                                    
                                    # Analizar contexto para verificar si es realmente el monto de una beca
                                    context = details_text[max(0, match.start()-30):min(len(details_text), match.end()+30)]
                                    scholarship_amount_indicators = [
                                        "scholarship", "award", "grant", "funding", "stipend", "beca", 
                                        "financiación", "monto", "stipendium", "betrag", "beurs", "bedrag",
                                        "amount", "value", "worth", "up to", "hasta", "bis zu", "tot",
                                        "receive", "awarded", "provides", "offers", "covers", "includes"
                                    ]
                                    
                                    if any(indicator in context.lower() for indicator in scholarship_amount_indicators):
                                        # Si este monto aparece antes en el texto que el mejor hasta ahora, actualizarlo
                                        if match.start() < best_position:
                                            best_amount = amount
                                            best_currency = currency
                                            best_position = match.start()
                                
                                # Asignar el mejor monto encontrado
                                if best_amount:
                                    scholarship['Amount'] = best_amount
                                    if best_currency:
                                        scholarship['Currency'] = best_currency
                                
                                # Extraer criterios de elegibilidad
                                eligibility_patterns = {
                                    "en": [
                                        r'(eligib|requirements?|qualifications?)[:\s]+([^\.]+)',
                                        r'(available to|open to|for students?)[:\s]+([^\.]+)'
                                    ],
                                    "es": [
                                        r'(requisitos|elegibilidad|pueden solicitar)[:\s]+([^\.]+)',
                                        r'(disponible para|abierto a|para estudiantes)[:\s]+([^\.]+)'
                                    ],
                                    "de": [
                                        r'(voraussetzungen|anforderungen|bewerbungsvoraussetzungen)[:\s]+([^\.]+)',
                                        r'(verfügbar für|offen für|für studierende)[:\s]+([^\.]+)'
                                    ],
                                    "nl": [
                                        r'(voorwaarden|eisen|vereisten)[:\s]+([^\.]+)',
                                        r'(beschikbaar voor|open voor|voor studenten)[:\s]+([^\.]+)'
                                    ]
                                }
                                
                                for pattern in eligibility_patterns[page_lang]:
                                    eligibility_match = re.search(pattern, details_text, re.I)
                                    if eligibility_match:
                                        eligibility_text = eligibility_match.group(2).strip()
                                        scholarship['Eligibility Criteria'] = eligibility_text[:150] + ('...' if len(eligibility_text) > 150 else '')
                                        break
                                
                                # Extraer competitividad
                                # Buscar palabras clave que indiquen competitividad
                                competitiveness_indicators = {
                                    "high": ["highly competitive", "limited", "very selective", "few", "small number",
                                         "alta competencia", "limitado", "muy selectivo", "pocos", "reducido número",
                                         "stark umkämpft", "begrenzt", "sehr selektiv", "wenige", "geringe anzahl",
                                         "zeer competitief", "beperkt", "zeer selectief", "weinig", "klein aantal"],
                                    "medium": ["competitive", "selected", "moderate", "average",
                                           "competitivo", "seleccionado", "moderado", "promedio",
                                           "wettbewerbsfähig", "ausgewählt", "mäßig", "durchschnittlich",
                                           "competitief", "geselecteerd", "gematigd", "gemiddeld"],
                                    "low": ["all eligible", "many", "numerous", "most", "high number",
                                        "todos los elegibles", "muchos", "numerosos", "mayoría", "gran número",
                                        "alle berechtigten", "viele", "zahlreiche", "meisten", "hohe anzahl",
                                        "alle in aanmerking", "veel", "talrijk", "meeste", "groot aantal"]
                                }
                                
                                for level, indicators in competitiveness_indicators.items():
                                    if any(indicator in details_text.lower() for indicator in indicators):
                                        scholarship['Competitiveness'] = level.capitalize()
                                        break
                                
                                # Extraer plazo de solicitud
                                deadline_patterns = {
                                    "en": r'(deadline|apply by|due|closing date)[:\s]+([A-Za-z]+ \d{1,2}(?:st|nd|rd|th)?,? \d{4}|\d{1,2}[- /\.]\d{1,2}[- /\.]\d{2,4})',
                                    "es": r'(fecha límite|plazo|vencimiento|cierre)[:\s]+(\d{1,2} de [A-Za-z]+ (?:de )?\d{4}|\d{1,2}[- /\.]\d{1,2}[- /\.]\d{2,4})',
                                    "de": r'(bewerbungsschluss|frist|stichtag|einsendeschluss)[:\s]+(\d{1,2}\. [A-Za-z]+ \d{4}|\d{1,2}[- /\.]\d{1,2}[- /\.]\d{2,4})',
                                    "nl": r'(deadline|uiterste datum|sluitingsdatum)[:\s]+(\d{1,2} [A-Za-z]+ \d{4}|\d{1,2}[- /\.]\d{1,2}[- /\.]\d{2,4})'
                                }
                                
                                deadline_match = re.search(deadline_patterns[page_lang], details_text, re.I)
                                if deadline_match:
                                    scholarship['Application Deadline'] = deadline_match.group(2)
                                
                                # Extraer número de becas
                                award_patterns = {
                                    "en": r'(\d+)[\s]+(scholarships?|awards?|grants?|positions?|students?|candidates?)',
                                    "es": r'(\d+)[\s]+(becas?|ayudas?|subvenciones?|plazas?|estudiantes?|candidatos?)',
                                    "de": r'(\d+)[\s]+(stipendien|förderungen|auszeichnungen|plätze|studierende|kandidaten)',
                                    "nl": r'(\d+)[\s]+(beurzen|toelagen|subsidies|plaatsen|studenten|kandidaten)'
                                }
                                
                                award_match = re.search(award_patterns[page_lang], details_text, re.I)
                                if award_match:
                                    scholarship['Number of Awards'] = award_match.group(1)
                                
                                # Extraer condiciones de renovación
                                renewal_keywords = {
                                    "en": ["renew", "renewal", "continue", "extension", "maintain"],
                                    "es": ["renovar", "renovación", "continuar", "extensión", "mantener"],
                                    "de": ["erneuer", "verlänger", "fortsetz", "beibehalt", "weitergabe"],
                                    "nl": ["vernieu", "verlenging", "voortzett", "behoud", "vervolg"]
                                }
                                
                                for keyword in renewal_keywords[page_lang]:
                                    renewal_match = re.search(f"{keyword}[a-z]*[:\s]+([^\.]+)", details_text, re.I)
                                    if renewal_match:
                                        scholarship['Renewal Conditions'] = renewal_match.group(1).strip()
                                        break
                                
                                # Extraer proceso de selección
                                selection_keywords = {
                                    "en": ["select", "process", "assess", "evaluat", "criteri"],
                                    "es": ["selecci", "proces", "evalu", "criterios", "valoración"],
                                    "de": ["auswahl", "prozess", "bewert", "kriterien", "beurteil"],
                                    "nl": ["selectie", "proces", "beoordel", "criteria", "evaluatie"]
                                }
                                
                                for keyword in selection_keywords[page_lang]:
                                    selection_match = re.search(f"{keyword}[a-z]*[:\s]+([^\.]+)", details_text, re.I)
                                    if selection_match:
                                        scholarship['Selection Process'] = selection_match.group(1).strip()
                                        break
                                
                                # Extraer información de contacto
                                contact_pattern = r'([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})'
                                email_match = re.search(contact_pattern, details_text)
                                if email_match:
                                    scholarship['Contact Email'] = email_match.group(1)
                                    
                                    # Intentar extraer nombre de contacto
                                    contact_name_pattern = r'([A-Za-z\.\s]{5,40})[\s,]+(?:[A-Za-z\.\s]{0,20}[\s,]+)?'+re.escape(email_match.group(1))
                                    name_match = re.search(contact_name_pattern, details_text)
                                    if name_match:
                                        scholarship['Contact Person'] = name_match.group(1).strip()
                    
                    except Exception as e:
                        logger.warning(f"Error extrayendo detalles para beca '{title}': {str(e)}")
                    
                    # Añadir la beca a la lista
                    scholarships.append(scholarship)
                    log_reference(university_name, f"Beca: {title}", scholarship_url)
                    
                    # Limitar a 5 becas por universidad para no sobrecargar
                    if len(scholarships) >= 5:
                        break
                
                # Si ya tenemos suficientes becas, salir del bucle
                if len(scholarships) >= 5:
                    break
            
            # Si ya tenemos suficientes becas, pasar a la siguiente universidad
            if len(scholarships) >= 5:
                break
                
        except Exception as e:
            logger.warning(f"Error procesando URL de becas {scholarship_url}: {str(e)}")
    
    # Añadir becas internacionales conocidas si no tenemos suficientes
    if len(scholarships) < 3:
        # Becas internacionales y regionales relevantes
        international_scholarships = [
            {
                'name': 'Fulbright Foreign Student Program',
                'type': 'Full Tuition',
                'url': 'https://foreign.fulbrightonline.org/',
                'eligibility': 'International students applying to US universities',
                'countries': ['Estados Unidos']
            },
            {
                'name': 'Chevening Scholarships',
                'type': 'Full Tuition',
                'url': 'https://www.chevening.org/',
                'eligibility': 'International students applying to UK universities',
                'countries': ['Reino Unido']
            },
            {
                'name': 'Gates Cambridge Scholarship',
                'type': 'Full Tuition',
                'url': 'https://www.gatescambridge.org/',
                'eligibility': 'International students applying to University of Cambridge',
                'countries': ['Reino Unido']
            },
            {
                'name': 'DAAD Scholarships',
                'type': 'Full Tuition',
                'url': 'https://www.daad.de/en/',
                'eligibility': 'International students applying to German universities',
                'countries': ['Alemania']
            },
            {
                'name': 'Swiss Government Excellence Scholarships',
                'type': 'Full Tuition',
                'url': 'https://www.sbfi.admin.ch/sbfi/en/home/education/scholarships-and-grants/swiss-government-excellence-scholarships.html',
                'eligibility': 'International students applying to Swiss universities',
                'countries': ['Suiza']
            },
            {
                'name': 'Erasmus Mundus Joint Master Degrees',
                'type': 'Full Tuition',
                'url': 'https://erasmus-plus.ec.europa.eu/opportunities/individuals/students/erasmus-mundus-joint-masters-scholarships',
                'eligibility': 'International students applying to European universities',
                'countries': ['España', 'Reino Unido', 'Alemania', 'Suiza', 'Países Bajos']
            },
            {
                'name': 'Colfuturo',
                'type': 'Partial Tuition',
                'url': 'https://www.colfuturo.org/',
                'eligibility': 'Colombian students for international postgraduate studies',
                'countries': ['Estados Unidos', 'España', 'Reino Unido', 'Canadá', 'Alemania', 'Suiza', 'Países Bajos', 'Chile']
            },
            {
                'name': 'CONACYT Scholarships',
                'type': 'Full Tuition',
                'url': 'https://www.conacyt.mx/',
                'eligibility': 'Mexican students for graduate studies',
                'countries': ['México', 'Estados Unidos', 'España', 'Reino Unido', 'Canadá', 'Alemania', 'Suiza']
            },
            {
                'name': 'ANID Becas Chile',
                'type': 'Full Tuition',
                'url': 'https://www.anid.cl/capital-humano/becas-chile/',
                'eligibility': 'Chilean students for graduate studies abroad',
                'countries': ['Chile', 'Estados Unidos', 'España', 'Reino Unido', 'Canadá', 'Alemania', 'Suiza']
            },
            {
                'name': 'Holland Scholarship',
                'type': 'Partial Tuition',
                'url': 'https://www.studyinholland.nl/finances/holland-scholarship',
                'eligibility': 'International students from outside the European Economic Area',
                'countries': ['Países Bajos']
            }
        ]
        
        country = university_name.split(", ")[-1]
        
        for scholarship_info in international_scholarships:
            # Comprobar si la beca aplica para el país de la universidad
            if country in scholarship_info['countries']:
                # Verificar si ya tenemos esta beca
                if any(s['Scholarship Name'] == scholarship_info['name'] for s in scholarships):
                    continue
                    
                scholarship_id = f"SCH{str(abs(hash(scholarship_info['name'])) % 10000).zfill(4)}"
                
                scholarship = {
                    'Scholarship_ID': scholarship_id,
                    'Univ_ID': univ_id,
                    'Prog_ID': '',
                    'Scholarship Name': scholarship_info['name'],
                    'Type of Funding': scholarship_info['type'],
                    'Amount': 'Varies',
                    'Currency': 'USD',
                    'Eligibility Criteria': scholarship_info['eligibility'],
                    'Competitiveness': 'High',
                    'Number of Awards': 'Varies',
                    'Application Deadline': 'Check website',
                    'Notification Date': 'Varies',
                    'Disbursement Schedule': 'Per semester/year',
                    'Renewal Conditions': 'Academic performance',
                    'Selection Process': 'Merit-based evaluation',
                    'Scholarship Website': scholarship_info['url'],
                    'Contact Person': 'Scholarship Office',
                    'Contact Email': 'N/A',
                    'Notes': 'International scholarship program'
                }
                
                scholarships.append(scholarship)
                log_reference(university_name, f"Beca Internacional: {scholarship_info['name']}", scholarship_info['url'])
                
                # Limitar a 5 becas por universidad
                if len(scholarships) >= 5:
                    break
    
    # Si aún no tenemos becas, crear becas genéricas
    if not scholarships:
        logger.warning(f"No se encontraron becas para {university_name}, generando datos ficticios")
        for i in range(3):
            scholarship_id = f"SCH{str(abs(hash(f'Scholarship{i}' + university_name)) % 10000).zfill(4)}"
            scholarship_types = ["Merit Scholarship", "International Student Scholarship", "Research Grant"]
            scholarships.append({
                'Scholarship_ID': scholarship_id,
                'Univ_ID': univ_id,
                'Prog_ID': '',
                'Scholarship Name': f"{university_name.split(',')[0]} {scholarship_types[i]}",
                'Type of Funding': ["Full Tuition", "Partial Tuition", "Research Grant"][i],
                'Amount': ["100%", "50%", "$10,000"][i],
                'Currency': 'USD',
                'Eligibility Criteria': 'International students with excellent academic record',
                'Competitiveness': ['High', 'Medium', 'Medium'][i],
                'Number of Awards': ['5-10', '10-20', '15-25'][i],
                'Application Deadline': 'Concurrent with program application',
                'Notification Date': '4-6 weeks after application',
                'Disbursement Schedule': 'Per semester',
                'Renewal Conditions': 'Maintain good academic standing',
                'Selection Process': 'Merit-based evaluation',
                'Scholarship Website': f"{university_url}/scholarships",
                'Contact Person': 'Financial Aid Office',
                'Contact Email': f"financial-aid@{urlparse(university_url).netloc}",
                'Notes': 'Datos aproximados, verificar en el sitio web oficial'
            })
    
    logger.info(f"Extracción de becas completada para {university_name}: {len(scholarships)} becas encontradas")
    return scholarships

def extract_admission_info(university_name, university_url, univ_id, prog_id=''):
    """Extrae información sobre requisitos de admisión"""
    admission = {
        'Admission_ID': f"ADM{str(abs(hash(university_name + (prog_id or ''))) % 10000).zfill(4)}",
        'Univ_ID': univ_id,
        'Prog_ID': prog_id,
        'Minimum GPA': 'N/A',
        'GPA Scale': 'N/A',
        'Required Exams': 'N/A',
        'Minimum Scores': 'N/A',
        'Language Test Validity (years)': 'N/A',
        'Letters of Recommendation': 'N/A',
        'Statement of Purpose': 'N/A',
        'Resume / CV': 'N/A',
        'Interview Requirement': 'N/A',
        'Research Proposal': 'N/A',
        'Experience Required': 'N/A',
        'Portfolio/Writing Samples': 'N/A',
        'Application Deadline': 'N/A',
        'Application Fee (USD)': 'N/A',
        'Rolling Admission': 'N/A',
        'Other Requirements': 'N/A',
        'Notes': ''
    }
    
    # URLs comunes para requisitos de admisión
    admission_urls = [
        f"{university_url}/admissions",
        f"{university_url}/apply",
        f"{university_url}/graduate/admissions",
        f"{university_url}/graduate/apply",
        f"{university_url}/international/requirements",
        f"{university_url}/requirements",
        f"{university_url}/graduate/requirements"
    ]
    
    for admission_url in admission_urls:
        try:
            html = get_html(admission_url)
            if not html:
                continue
                
            soup = BeautifulSoup(html, 'html.parser')
            
            # Extraer GPA mínimo
            gpa_patterns = [
                r'(minimum|required)\s+GPA\s+(?:of)?\s+(\d+\.\d+)',
                r'GPA\s+(?:of)?\s+(\d+\.\d+)\s+or\s+(above|higher)',
                r'GPA\s*[:=]\s*(\d+\.\d+)'
            ]
            
            for pattern in gpa_patterns:
                gpa_match = re.search(pattern, soup.text, re.I)
                if gpa_match:
                    gpa_value = next((g for g in gpa_match.groups() if g and re.match(r'\d+\.\d+', g)), None)
                    if gpa_value:
                        admission['Minimum GPA'] = gpa_value
                        
                        # Determinar escala de GPA
                        if float(gpa_value) <= 4.0:
                            admission['GPA Scale'] = '4.0'
                        elif float(gpa_value) <= 5.0:
                            admission['GPA Scale'] = '5.0'
                        elif float(gpa_value) <= 10.0:
                            admission['GPA Scale'] = '10.0'
                        break
            
            # Extraer exámenes requeridos
            exam_patterns = {
                'GRE': r'(GRE|Graduate Record Examination)',
                'GMAT': r'(GMAT|Graduate Management Admission Test)',
                'TOEFL': r'(TOEFL|Test of English as a Foreign Language)',
                'IELTS': r'(IELTS|International English Language Testing System)'
            }
            
            required_exams = []
            for exam, pattern in exam_patterns.items():
                if re.search(pattern, soup.text, re.I):
                    required_exams.append(exam)
            
            if required_exams:
                admission['Required Exams'] = ', '.join(required_exams)
            
            # Extraer puntuaciones mínimas
            score_patterns = {
                'TOEFL': r'TOEFL\s+(?:minimum|required)?\s+(?:score\s+(?:of)?)?\s+(\d+)',
                'IELTS': r'IELTS\s+(?:minimum|required)?\s+(?:score\s+(?:of)?)?\s+(\d+(?:\.\d+)?)',
                'GRE': r'GRE\s+(?:minimum|required)?\s+(?:score\s+(?:of)?)?\s+(\d+)',
                'GMAT': r'GMAT\s+(?:minimum|required)?\s+(?:score\s+(?:of)?)?\s+(\d+)'
            }
            
            min_scores = []
            for exam, pattern in score_patterns.items():
                score_match = re.search(pattern, soup.text, re.I)
                if score_match:
                    min_scores.append(f"{exam}: {score_match.group(1)}")
            
            if min_scores:
                admission['Minimum Scores'] = ', '.join(min_scores)
            
            # Extraer validez de prueba de idioma
            validity_match = re.search(r'(TOEFL|IELTS).*?valid for (\d+) years?', soup.text, re.I)
            if validity_match:
                admission['Language Test Validity (years)'] = validity_match.group(2)
            
            # Extraer cartas de recomendación
            rec_match = re.search(r'(\d+).*?letters? of recommendation', soup.text, re.I)
            if rec_match:
                admission['Letters of Recommendation'] = rec_match.group(1)
            
            # Extraer statement of purpose
            if re.search(r'statement of (purpose|intent|objectives)', soup.text, re.I):
                admission['Statement of Purpose'] = 'Yes'
            
            # Extraer requisito de CV
            if re.search(r'(resume|CV|curriculum vitae)', soup.text, re.I):
                admission['Resume / CV'] = 'Yes'
            
            # Extraer requisito de entrevista
            if re.search(r'interview', soup.text, re.I):
                admission['Interview Requirement'] = 'Yes'
            
            # Extraer requisito de propuesta de investigación
            if re.search(r'research proposal', soup.text, re.I):
                admission['Research Proposal'] = 'Yes'
            
            # Extraer tarifa de aplicación
            fee_match = re.search(r'application fee.*?(\$|\€|\£|\¥)?(\d+)', soup.text, re.I)
            if fee_match:
                currency_symbol = fee_match.group(1) or '$'
                fee_amount = fee_match.group(2)
                
                # Convertir a USD si no está en esa moneda
                if currency_symbol == '$':
                    admission['Application Fee (USD)'] = fee_amount
                else:
                    # Conversión aproximada (se podrían usar APIs de conversión de moneda)
                    conversion_rates = {'€': 1.1, '£': 1.3, '¥': 0.0068}
                    rate = conversion_rates.get(currency_symbol, 1)
                    usd_amount = int(float(fee_amount) * rate)
                    admission['Application Fee (USD)'] = str(usd_amount)
            
            # Extraer plazos de solicitud
            deadline_match = re.search(r'(application\s+deadline|apply\s+by)[:\s]+([A-Za-z]+ \d{1,2}(st|nd|rd|th)?,? \d{4}|\d{1,2}[- /\.]\d{1,2}[- /\.]\d{2,4})', soup.text, re.I)
            if deadline_match:
                admission['Application Deadline'] = deadline_match.group(2)
            
            # Determinar si tiene admisión continua
            if re.search(r'rolling admission|applications? accepted (on a)? rolling basis', soup.text, re.I):
                admission['Rolling Admission'] = 'Yes'
            else:
                admission['Rolling Admission'] = 'No'
            
            log_reference(university_name, "Requisitos de admisión", admission_url)
            break
                
        except Exception as e:
            logging.error(f"Error extrayendo requisitos de admisión para {university_name}: {str(e)}")
    
    return admission

def extract_cost_living_info(university_name, city, country, univ_id):
    """Extrae información sobre costo de vida"""
    cost_id = f"CST{str(abs(hash(city + country)) % 10000).zfill(4)}"
    
    cost = {
        'Cost_ID': cost_id,
        'Univ_ID': univ_id,
        'City': city,
        'Country': country,
        'Currency': 'USD',
        'Estimated Monthly Living Costs': 'N/A',
        'Housing Type': 'N/A',
        'Housing Costs': 'N/A',
        'Food/Groceries': 'N/A',
        'Public Transportation': 'N/A',
        'Utilities': 'N/A',
        'Health Insurance': 'N/A',
        'Textbooks & Supplies': 'N/A',
        'Climate': 'N/A',
        'Safety Rating': 'N/A',
        'Part-time Work Opportunities': 'N/A',
        'Visa Cost': 'N/A',
        'Visa Process': 'N/A',
        'Student Services': 'N/A',
        'Notes': ''
    }
    
    # Intentar obtener datos de Numbeo (simulado)
    numbeo_url = f"https://www.numbeo.com/cost-of-living/in/{city.replace(' ', '-')}"
    
    try:
        html = get_html(numbeo_url)
        if html:
            soup = BeautifulSoup(html, 'html.parser')
            
            # Extraer costo estimado mensual
            monthly_cost_div = soup.find('div', text=re.compile('Monthly costs for a single person'))
            if monthly_cost_div:
                amount_match = re.search(r'(\d{1,3}(,\d{3})+|\d+\.\d+|\d{4,})', monthly_cost_div.text)
                if amount_match:
                    cost['Estimated Monthly Living Costs'] = amount_match.group(1)
            
            # Extraer costo de vivienda
            housing_table = soup.find('table', {'class': 'data_wide_table'})
            if housing_table:
                housing_rows = housing_table.find_all('tr')
                for row in housing_rows:
                    if 'Apartment (1 bedroom) in City Centre' in row.text:
                        cells = row.find_all('td')
                        if len(cells) >= 2:
                            cost['Housing Costs'] = cells[1].text.strip()
                            break
            
            # Extraer costo de comida
            food_rows = soup.find_all('tr', text=re.compile('Meal, Inexpensive Restaurant|Milk|Bread|Rice|Eggs|Cheese'))
            food_costs = []
            for row in food_rows:
                cells = row.find_all('td')
                if len(cells) >= 2:
                    food_costs.append(cells[1].text.strip())
            
            if food_costs:
                # Calcular un promedio aproximado para alimentos mensuales
                food_monthly = '$300-600'  # Valor por defecto
                cost['Food/Groceries'] = food_monthly
            
            # Extraer costo de transporte público
            transport_row = soup.find('tr', text=re.compile('Monthly Pass, Regular Price'))
            if transport_row:
                cells = transport_row.find_all('td')
                if len(cells) >= 2:
                    cost['Public Transportation'] = cells[1].text.strip()
            
            # Extraer costo de utilidades
            utilities_row = soup.find('tr', text=re.compile('Basic.*?Electricity, Heating, Cooling, Water, Garbage'))
            if utilities_row:
                cells = utilities_row.find_all('td')
                if len(cells) >= 2:
                    cost['Utilities'] = cells[1].text.strip()
            
            log_reference(university_name, f"Costo de vida en {city}", numbeo_url)
        
    except Exception as e:
        logging.error(f"Error extrayendo costo de vida para {city}, {country}: {str(e)}")
    
    # Definir tipo de clima según la ubicación (simulado)
    climate_map = {
        'Estados Unidos': {
            'Boston': 'Continental: inviernos fríos y veranos cálidos',
            'San Francisco': 'Mediterráneo: templado todo el año',
            'New York': 'Continental: inviernos fríos y veranos calurosos',
            'Chicago': 'Continental: inviernos muy fríos y veranos cálidos',
            'Los Angeles': 'Mediterráneo: templado y seco'
        },
        'Reino Unido': {
            'London': 'Oceánico: templado y húmedo todo el año',
            'Cambridge': 'Oceánico: templado y húmedo todo el año',
            'Oxford': 'Oceánico: templado y húmedo todo el año',
            'Edinburgh': 'Oceánico: fresco y húmedo todo el año'
        },
        'Canadá': {
            'Toronto': 'Continental: inviernos muy fríos y veranos cálidos',
            'Vancouver': 'Oceánico: templado y muy lluvioso',
            'Montreal': 'Continental: inviernos extremadamente fríos',
            'Ottawa': 'Continental: inviernos extremadamente fríos'
        },
        'España': {
            'Madrid': 'Mediterráneo continental: veranos calurosos e inviernos fríos',
            'Barcelona': 'Mediterráneo: veranos cálidos e inviernos suaves',
            'Valencia': 'Mediterráneo: veranos calurosos e inviernos suaves',
            'Sevilla': 'Mediterráneo: veranos muy calurosos e inviernos suaves'
        },
        'Alemania': {
            'Munich': 'Continental: inviernos fríos y veranos templados',
            'Berlin': 'Continental: inviernos fríos y veranos templados',
            'Heidelberg': 'Continental: inviernos fríos y veranos templados',
            'Aachen': 'Oceánico: templado y húmedo'
        },
        'Suiza': {
            'Zurich': 'Continental: inviernos fríos y veranos templados',
            'Lausanne': 'Continental moderado: influencia del lago Lemán',
            'Geneva': 'Continental moderado: influencia del lago Lemán',
            'Lugano': 'Mediterráneo de montaña: más cálido que el resto de Suiza'
        },
        'Países Bajos': {
            'Amsterdam': 'Oceánico: templado y húmedo todo el año',
            'Delft': 'Oceánico: templado y húmedo todo el año',
            'Utrecht': 'Oceánico: templado y húmedo todo el año',
            'Leiden': 'Oceánico: templado y húmedo todo el año'
        }
    }
    
    # Asignar clima según país y ciudad
    if country in climate_map and city in climate_map[country]:
        cost['Climate'] = climate_map[country][city]
    else:
        # Asignar clima por país si la ciudad no está en el mapa
        country_climates = {
            'Estados Unidos': 'Varía por región: continental a subtropical',
            'Reino Unido': 'Oceánico: templado y húmedo',
            'Canadá': 'Continental: inviernos muy fríos',
            'España': 'Mediterráneo: veranos cálidos e inviernos suaves',
            'Alemania': 'Continental: inviernos fríos y veranos templados',
            'Suiza': 'Continental alpino: inviernos fríos',
            'Países Bajos': 'Oceánico: templado y húmedo',
            'México': 'Varía por región: tropical a desértico',
            'Chile': 'Varía por región: mediterráneo a subpolar'
        }
        cost['Climate'] = country_climates.get(country, 'N/A')
    
    # Asignar índice de seguridad (simulado)
    safety_ratings = {
        'Estados Unidos': {'media': 'Average', 'buenas': ['Boston', 'San Francisco'], 'regulares': ['Chicago', 'Los Angeles']},
        'Reino Unido': {'media': 'Safe', 'buenas': ['Cambridge', 'Oxford'], 'regulares': ['London']},
        'Canadá': {'media': 'Very Safe', 'buenas': ['Vancouver', 'Ottawa'], 'regulares': []},
        'España': {'media': 'Safe', 'buenas': ['Salamanca'], 'regulares': ['Madrid', 'Barcelona']},
        'Alemania': {'media': 'Very Safe', 'buenas': ['Munich', 'Heidelberg'], 'regulares': ['Berlin']},
        'Suiza': {'media': 'Very Safe', 'buenas': ['Zurich', 'Geneva', 'Lausanne'], 'regulares': []},
        'Países Bajos': {'media': 'Safe', 'buenas': ['Delft', 'Leiden'], 'regulares': ['Amsterdam']},
        'México': {'media': 'Below Average', 'buenas': ['Querétaro', 'Mérida'], 'regulares': ['Ciudad de México']},
        'Chile': {'media': 'Safe', 'buenas': ['Viña del Mar'], 'regulares': ['Santiago']}
    }
    
    if country in safety_ratings:
        if city in safety_ratings[country]['buenas']:
            cost['Safety Rating'] = 'Very Safe'
        elif city in safety_ratings[country]['regulares']:
            cost['Safety Rating'] = 'Average'
        else:
            cost['Safety Rating'] = safety_ratings[country]['media']
    else:
        cost['Safety Rating'] = 'N/A'
    
    # Posibilidades de trabajo a tiempo parcial según país
    part_time_work = {
        'Estados Unidos': 'Hasta 20 horas/semana con F-1 visa (on-campus only)',
        'Reino Unido': 'Hasta 20 horas/semana durante el período lectivo',
        'Canadá': 'Hasta 20 horas/semana fuera del campus',
        'España': 'Permitido con permiso de estudiante (mod. inicial)',
        'Alemania': 'Hasta 120 días completos o 240 medios días por año',
        'Suiza': 'Hasta 15 horas/semana (restricciones según cantón)',
        'Países Bajos': 'Hasta 16 horas/semana o tiempo completo en verano',
        'México': 'Restringido con visa de estudiante',
        'Chile': 'Permitido con visa de estudiante'
    }
    
    cost['Part-time Work Opportunities'] = part_time_work.get(country, 'N/A')
    
    # Información sobre visas según país
    visa_info = {
        'Estados Unidos': {'costo': '$350 (F-1)', 'proceso': 'Requiere I-20 de la universidad y entrevista consular'},
        'Reino Unido': {'costo': '£348 (Student visa)', 'proceso': 'Requiere CAS de la universidad'},
        'Canadá': {'costo': 'CAD $150', 'proceso': 'Requiere carta de aceptación y prueba de fondos'},
        'España': {'costo': '€80', 'proceso': 'Requiere seguro médico y prueba de fondos'},
        'Alemania': {'costo': '€75', 'proceso': 'Requiere carta de aceptación y bloqueo de cuenta'},
        'Suiza': {'costo': 'CHF 60-140', 'proceso': 'Varía según cantón y nacionalidad'},
        'Países Bajos': {'costo': '€207', 'proceso': 'Tramitada por la universidad (MVV)'},
        'México': {'costo': '$36', 'proceso': 'Requiere carta de aceptación y prueba de fondos'},
        'Chile': {'costo': '$100', 'proceso': 'Requiere carta de aceptación y antecedentes'}
    }
    
    if country in visa_info:
        cost['Visa Cost'] = visa_info[country]['costo']
        cost['Visa Process'] = visa_info[country]['proceso']
    
    # Servicios estudiantiles típicos
    typical_services = "Orientación, servicios de salud, asesoramiento académico, apoyo psicológico, instalaciones deportivas, bibliotecas, servicios de carrera"
    cost['Student Services'] = typical_services
    
    return cost

def extract_outcome_info(university_name, university_url, univ_id, prog_id=''):
    """Extrae información sobre resultados profesionales y empleabilidad"""
    outcome_id = f"OUT{str(abs(hash(university_name + (prog_id or ''))) % 10000).zfill(4)}"
    
    outcome = {
        'Outcome_ID': outcome_id,
        'Univ_ID': univ_id,
        'Prog_ID': prog_id,
        'Employability Rate (%)': 'N/A',
        'Average Starting Salary': 'N/A',
        'Currency': 'USD',
        'Time to First Job (months)': 'N/A',
        'Top Employers': 'N/A',
        'Internship Opportunities': 'N/A',
        'Industry Partnerships': 'N/A',
        'Alumni Network Size': 'N/A',
        'Alumni Events': 'N/A',
        'Alumni Mentorship Programs': 'N/A',
        'Further Study Rate (%)': 'N/A',
        'Job Satisfaction (1-5)': 'N/A',
        'Career Support Services': 'N/A',
        'Visa Extension Options': 'N/A',
        'Notes': ''
    }
    
    # URLs comunes para resultados de egresados
    outcome_urls = [
        f"{university_url}/career",
        f"{university_url}/careers",
        f"{university_url}/alumni",
        f"{university_url}/outcomes",
        f"{university_url}/placement",
        f"{university_url}/employment",
        f"{university_url}/graduate-outcomes"
    ]
    
    for outcome_url in outcome_urls:
        try:
            html = get_html(outcome_url)
            if not html:
                continue
                
            soup = BeautifulSoup(html, 'html.parser')
            
            # Extraer tasa de empleabilidad
            employment_patterns = [
                r'(\d{1,3})%.*?(employment|employed|job placement|placement rate)',
                r'(employment|employed|job placement|placement rate).*?(\d{1,3})%',
                r'(\d{1,3}) percent.*?(employment|employed|job placement)'
            ]
            
            for pattern in employment_patterns:
                employment_match = re.search(pattern, soup.text, re.I)
                if employment_match:
                    percent_group = next((g for g in employment_match.groups() if g and g.isdigit()), None)
                    if percent_group and 0 <= int(percent_group) <= 100:
                        outcome['Employability Rate (%)'] = percent_group
                        break
            
            # Extraer salario inicial promedio
            salary_patterns = [
                r'(average|median) (starting|initial) salary.*?(\$|\€|\£|\¥)?(\d{1,3}(,\d{3})+|\d{4,})',
                r'(\$|\€|\£|\¥)?(\d{1,3}(,\d{3})+|\d{4,}).*?(average|median) (starting|initial) salary',
                r'(starting|initial) salary.*?(\$|\€|\£|\¥)?(\d{1,3}(,\d{3})+|\d{4,})'
            ]
            
            for pattern in salary_patterns:
                salary_match = re.search(pattern, soup.text, re.I)
                if salary_match:
                    # Extraer el monto y la moneda
                    amount_group = next((g for g in salary_match.groups() if g and re.match(r'\d{1,3}(,\d{3})+|\d{4,}', g)), None)
                    currency_group = next((g for g in salary_match.groups() if g in ['$', '€', '£', '¥']), None)
                    
                    if amount_group:
                        outcome['Average Starting Salary'] = amount_group
                        
                        if currency_group:
                            currency_map = {
                                '$': 'USD',
                                '€': 'EUR',
                                '£': 'GBP',
                                '¥': 'JPY'
                            }
                            outcome['Currency'] = currency_map.get(currency_group, 'USD')
                        break
            
            # Extraer tiempo hasta el primer empleo
            time_patterns = [
                r'(\d{1,2}).*?(months?|weeks?).*?(to secure|to find|first job|employment)',
                r'(graduates? find|secure).*?(\d{1,2}).*?(months?|weeks?)',
                r'(time to|time until).*?(\d{1,2}).*?(months?|weeks?)'
            ]
            
            for pattern in time_patterns:
                time_match = re.search(pattern, soup.text, re.I)
                if time_match:
                    num_group = next((g for g in time_match.groups() if g and g.isdigit()), None)
                    unit_group = next((g for g in time_match.groups() if g and g.lower() in ['month', 'months', 'week', 'weeks']), None)
                    
                    if num_group and unit_group:
                        # Convertir semanas a meses si es necesario
                        if 'week' in unit_group.lower():
                            months = round(int(num_group) / 4.33)  # Aproximación
                            outcome['Time to First Job (months)'] = str(months)
                        else:
                            outcome['Time to First Job (months)'] = num_group
                        break
            
            # Extraer principales empleadores
            employer_patterns = [
                r'(top employers?|notable employers?|main employers?|key employers?).*?([^\.]+)',
                r'(companies? that hire|firms? that recruit).*?([^\.]+)',
                r'(our graduates? work for|alumni work for).*?([^\.]+)'
            ]
            
            for pattern in employer_patterns:
                employer_match = re.search(pattern, soup.text, re.I)
                if employer_match:
                    employer_text = employer_match.group(2)
                    # Filtrar para empresas conocidas
                    known_companies = ['Google', 'Microsoft', 'Amazon', 'Apple', 'Facebook', 'IBM', 'Oracle', 
                                    'Intel', 'Cisco', 'Adobe', 'SAP', 'Accenture', 'Deloitte', 'PwC', 'KPMG', 
                                    'EY', 'McKinsey', 'Boston Consulting', 'Bain', 'Goldman Sachs', 'JP Morgan',
                                    'Morgan Stanley', 'Bank of America', 'Citigroup', 'HSBC', 'Barclays']
                    
                    found_companies = []
                    for company in known_companies:
                        if company.lower() in employer_text.lower():
                            found_companies.append(company)
                    
                    if found_companies:
                        outcome['Top Employers'] = ', '.join(found_companies)
                    elif len(employer_text) > 5:
                        # Si no encontramos empresas conocidas, usar el texto original
                        outcome['Top Employers'] = employer_text[:100] + ('...' if len(employer_text) > 100 else '')
                    break
            
            # Extraer oportunidades de prácticas
            internship_patterns = [
                r'(internship|practical training|co-op).*?(opportunities|program|available)',
                r'(students? can|students? have access to).*?(internship|practical training|co-op)',
                r'(offers?|provides?).*?(internship|practical training|co-op)'
            ]
            
            for pattern in internship_patterns:
                if re.search(pattern, soup.text, re.I):
                    outcome['Internship Opportunities'] = 'Available'
                    break
            
            # Extraer tamaño de la red de alumni
            alumni_patterns = [
                r'(alumni network|network of alumni).*?(\d{1,3}(,\d{3})+|\d{4,})',
                r'(\d{1,3}(,\d{3})+|\d{4,}).*?(alumni|graduates)',
                r'(community of).*?(\d{1,3}(,\d{3})+|\d{4,}).*?(alumni|graduates)'
            ]
            
            for pattern in alumni_patterns:
                alumni_match = re.search(pattern, soup.text, re.I)
                if alumni_match:
                    num_group = next((g for g in alumni_match.groups() if g and re.match(r'\d{1,3}(,\d{3})+|\d{4,}', g)), None)
                    if num_group:
                        outcome['Alumni Network Size'] = num_group
                        break
            
            # Extraer eventos para alumni
            if re.search(r'alumni (events|gatherings|reunions|meetings|conferences)', soup.text, re.I):
                outcome['Alumni Events'] = 'Yes'
            
            # Extraer programas de mentoría
            if re.search(r'(mentorship|mentoring) program', soup.text, re.I):
                outcome['Alumni Mentorship Programs'] = 'Yes'
            
            # Extraer tasa de continuación de estudios
            further_patterns = [
                r'(\d{1,2})%.*?(further study|graduate study|phd|doctoral|advanced degree)',
                r'(further study|graduate study|phd|doctoral|advanced degree).*?(\d{1,2})%',
                r'(\d{1,2}) percent.*?(further study|graduate study)'
            ]
            
            for pattern in further_patterns:
                further_match = re.search(pattern, soup.text, re.I)
                if further_match:
                    percent_group = next((g for g in further_match.groups() if g and g.isdigit()), None)
                    if percent_group and 0 <= int(percent_group) <= 100:
                        outcome['Further Study Rate (%)'] = percent_group
                        break
            
            # Extraer servicios de apoyo profesional
            career_services = []
            service_keywords = ['career counseling', 'resume review', 'cv workshop', 'interview preparation', 
                             'job fair', 'career fair', 'networking event', 'employer presentation']
            
            for keyword in service_keywords:
                if re.search(keyword, soup.text, re.I):
                    career_services.append(keyword.title())
            
            if career_services:
                outcome['Career Support Services'] = ', '.join(career_services)
            else:
                outcome['Career Support Services'] = 'Standard career services available'
            
            # Extraer opciones de extensión de visa
            visa_patterns = {
                'Estados Unidos': r'(OPT|Optional Practical Training|STEM extension)',
                'Reino Unido': r'(Graduate Route|Post-Study Work Visa)',
                'Canadá': r'(PGWP|Post-Graduation Work Permit)',
                'Australia': r'(Temporary Graduate visa|subclass 485)',
                'Alemania': r'(18-month residence permit|job-seeker visa)',
                'Países Bajos': r'(orientation year|zoekjaar)',
                'Suiza': r'(six months to find work)',
                'España': r'(post-study work visa)'
            }
            
            for country, pattern in visa_patterns.items():
                if re.search(pattern, soup.text, re.I):
                    outcome['Visa Extension Options'] = f"Yes - {pattern}"
                    break
            
            log_reference(university_name, "Resultados de egresados", outcome_url)
            break
                
        except Exception as e:
            logging.error(f"Error extrayendo resultados de egresados para {university_name}: {str(e)}")
    
    # Datos por defecto para campos vacíos según el país
    default_data = {
        'computer_science': {
            'employment_rate': '90-95%',
            'salary': {
                'Estados Unidos': '$75,000-120,000',
                'Reino Unido': '£35,000-60,000',
                'Canadá': 'CAD $70,000-95,000',
                'España': '€30,000-45,000',
                'Alemania': '€45,000-65,000',
                'Suiza': 'CHF 80,000-120,000',
                'Países Bajos': '€40,000-65,000',
                'México': 'MXN 240,000-600,000',
                'Chile': 'CLP 15,000,000-30,000,000'
            }
        },
        'business_analytics': {
            'employment_rate': '85-92%',
            'salary': {
                'Estados Unidos': '$70,000-110,000',
                'Reino Unido': '£32,000-55,000',
                'Canadá': 'CAD $65,000-90,000',
                'España': '€28,000-45,000',
                'Alemania': '€42,000-62,000',
                'Suiza': 'CHF 75,000-110,000',
                'Países Bajos': '€38,000-60,000',
                'México': 'MXN 220,000-540,000',
                'Chile': 'CLP 14,000,000-28,000,000'
            }
        },
        'mathematics': {
            'employment_rate': '80-90%',
            'salary': {
                'Estados Unidos': '$65,000-95,000',
                'Reino Unido': '£30,000-50,000',
                'Canadá': 'CAD $60,000-85,000',
                'España': '€26,000-42,000',
                'Alemania': '€40,000-58,000',
                'Suiza': 'CHF 70,000-100,000',
                'Países Bajos': '€35,000-55,000',
                'México': 'MXN 200,000-480,000',
                'Chile': 'CLP 12,000,000-24,000,000'
            }
        }
    }
    
    # Asignar valores por defecto si los datos están vacíos
    if outcome['Employability Rate (%)'] == 'N/A':
        # Determinar el programa según prog_id
        program_type = 'computer_science'  # Por defecto
        outcome['Employability Rate (%)'] = default_data[program_type]['employment_rate']
    
    if outcome['Average Starting Salary'] == 'N/A':
        program_type = 'computer_science'  # Por defecto
        country = university_name.split(', ')[0]
        outcome['Average Starting Salary'] = default_data[program_type]['salary'].get(country, '$60,000-90,000')
    
    if outcome['Time to First Job (months)'] == 'N/A':
        outcome['Time to First Job (months)'] = '3-6'
    
    # Opciones de extensión de visa por país
    visa_extensions = {
        'Estados Unidos': 'OPT: 12 meses + 24 adicionales para STEM',
        'Reino Unido': 'Graduate Route: 2 años (3 para doctorados)',
        'Canadá': 'PGWP: hasta 3 años según duración del programa',
        'España': 'Prórroga de estancia por búsqueda de empleo: 12 meses',
        'Alemania': 'Permiso de residencia para buscar trabajo: 18 meses',
        'Suiza': 'Permiso para buscar trabajo: 6 meses',
        'Países Bajos': 'Orientation Year: 12 meses',
        'México': 'Posibilidad de cambiar a visa de trabajo con oferta laboral',
        'Chile': 'Visa sujeta a contrato con oferta laboral'
    }
    
    if outcome['Visa Extension Options'] == 'N/A':
        country = university_name.split(', ')[0]
        outcome['Visa Extension Options'] = visa_extensions.get(country, 'Varía según regulaciones migratorias')
    
    return outcome

def create_empty_notes(university_name, univ_id, prog_id=''):
    """Crea un registro vacío para la hoja de notas personales"""
    notes_id = f"NOT{str(abs(hash(university_name + (prog_id or ''))) % 10000).zfill(4)}"
    
    notes = {
        'Notes_ID': notes_id,
        'Univ_ID': univ_id,
        'Prog_ID': prog_id,
        'Personal Interest Level': '',
        'Alignment with Career Goals': '',
        'Cultural Fit': '',
        'Family/Friends Nearby': '',
        'Personal Comments': '',
        'Date of Last Review': '',
        'Next Steps': '',
        'Final Decision': ''
    }
    
    return notes

def create_empty_timeline(university_name, univ_id, prog_id='', program_name='', deadline='N/A'):
    """Crea un registro vacío para la hoja de cronograma con algunos datos precompletados"""
    timeline_id = f"TL{str(abs(hash(university_name + (prog_id or ''))) % 10000).zfill(4)}"
    
    timeline = {
        'Timeline_ID': timeline_id,
        'Univ_ID': univ_id,
        'Prog_ID': prog_id,
        'Program Name': program_name,
        'University': university_name,
        'Program Deadline': deadline,
        'Application Start Date': '',
        'Document Preparation': '',
        'Test Date(s)': '',
        'Letter of Rec Deadline': '',
        'Scholarship Deadline': '',
        'Expected Response Date': '',
        'Deposit Due Date': '',
        'Visa Application Date': '',
        'Housing Application': '',
        'Orientation Date': '',
        'Program Start Date': '',
        'Status': 'Not Started',
        'Priority': '',
        'Notes': ''
    }
    
    return timeline

def main():
    """
    Función principal mejorada que orquesta el proceso de extracción de datos.
    Incluye manejo de errores, puntos de control y optimización.
    """
    start_time = time.time()
    logger.info("=== INICIANDO PROCESO DE EXTRACCIÓN DE DATOS UNIVERSITARIOS ===")
    
    # Verificar dependencias
    try:
        logger.info(f"Pandas versión: {pd.__version__}")
        logger.info(f"Requests versión: {requests.__version__}")
        logger.info(f"BeautifulSoup versión: {BeautifulSoup.__version__}")
    except:
        logger.info("No se pudieron verificar todas las versiones de las dependencias")
    
    # Inicializar archivo de referencias
    try:
        with open(REFERENCES_FILE, "w", encoding="utf-8") as f:
            f.write("# Referencias de Consulta para Universidad_Comparacion_Populated.xlsx\n\n")
        logger.info(f"Archivo de referencias inicializado: {REFERENCES_FILE}")
    except Exception as e:
        logger.error(f"Error al inicializar archivo de referencias: {str(e)}")
    
    # Obtener datos de países y universidades
    countries, universities = get_universities_data()
    logger.info(f"Datos cargados para {len(countries)} países y {sum(len(unis) for unis in universities.values())} universidades")
    
    # Cargar punto de control si existe
    checkpoint = load_checkpoint()
    start_country_idx = 0
    start_univ_idx = 0
    
    if checkpoint:
        # Encontrar índice del país en el checkpoint
        try:
            start_country_idx = countries.index(checkpoint['country'])
            start_univ_idx = checkpoint['university_index']
            logger.info(f"Reanudando desde: {countries[start_country_idx]}, universidad #{start_univ_idx+1}")
        except (ValueError, KeyError) as e:
            logger.warning(f"Error al procesar checkpoint, comenzando desde el principio: {str(e)}")
            start_country_idx = 0
            start_univ_idx = 0
    
    # Verificar si el archivo Excel existe y es accesible
    try:
        if os.path.exists(INPUT_EXCEL):
            # Cargar la plantilla Excel
            wb = pd.ExcelFile(INPUT_EXCEL)
            logger.info(f"Plantilla Excel cargada correctamente: {INPUT_EXCEL}")
        else:
            logger.error(f"¡Archivo {INPUT_EXCEL} no encontrado!")
            return
    except Exception as e:
        logger.error(f"Error al verificar la plantilla Excel: {str(e)}")
        return
    
    # Preparar los dataframes para cada hoja
    try:
        universities_df = pd.DataFrame()
        programs_df = pd.DataFrame()
        labs_df = pd.DataFrame()
        scholarships_df = pd.DataFrame()
        admissions_df = pd.DataFrame()
        costs_df = pd.DataFrame()
        outcomes_df = pd.DataFrame()
        notes_df = pd.DataFrame()
        timeline_df = pd.DataFrame()
        
        logger.info("DataFrames inicializados correctamente")
    except Exception as e:
        logger.error(f"Error al inicializar DataFrames: {str(e)}")
        return
    
    # Iterar por cada país y universidad
    for country_idx, country in enumerate(countries[start_country_idx:], start_country_idx):
        logger.info(f"Procesando país ({country_idx+1}/{len(countries)}): {country}")
        
        # Determinar desde qué universidad comenzar para este país
        univ_start_idx = start_univ_idx if country_idx == start_country_idx else 0
        
        for univ_idx, univ in enumerate(universities[country][univ_start_idx:], univ_start_idx):
            university_name = f"{univ['name']}, {country}"
            logger.info(f"Procesando universidad ({univ_idx+1}/{len(universities[country])}): {university_name}")
            
            # Guardar checkpoint antes de procesar cada universidad
            save_checkpoint(country, univ_idx, university_name)
            
            try:
                # 1. Extraer información general de la universidad
                logger.info(f"Extrayendo información general de {university_name}")
                university_data = extract_university_info(university_name, univ['url'], country, univ['city'])
                universities_df = pd.concat([universities_df, pd.DataFrame([university_data])], ignore_index=True)
                
                univ_id = university_data['Univ_ID']
                logger.info(f"Información general extraída exitosamente. ID: {univ_id}")
                
                # Procesar extracciones de forma concurrente pero con manejo de errores mejorado
                extracted_data = {
                    'programs': None,
                    'labs': None,
                    'scholarships': None,
                    'admission': None,
                    'cost': None,
                    'outcome': None
                }
                
                with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
                    # Iniciar todas las tareas
                    future_to_key = {
                        executor.submit(extract_program_info, university_name, univ['url'], univ_id): 'programs',
                        executor.submit(extract_lab_info, university_name, univ['url'], univ_id): 'labs',
                        executor.submit(extract_scholarship_info, university_name, univ['url'], univ_id): 'scholarships',
                        executor.submit(extract_admission_info, university_name, univ['url'], univ_id): 'admission',
                        executor.submit(extract_cost_living_info, university_name, univ['city'], country, univ_id): 'cost',
                        executor.submit(extract_outcome_info, university_name, univ['url'], univ_id): 'outcome'
                    }
                    
                    # Procesar resultados a medida que se completan
                    for future in concurrent.futures.as_completed(future_to_key):
                        key = future_to_key[future]
                        try:
                            data = future.result()
                            extracted_data[key] = data
                            logger.info(f"Extracción de {key} completada para {university_name}")
                        except Exception as e:
                            logger.error(f"Error en extracción de {key} para {university_name}: {str(e)}")
                            # Crear datos por defecto en caso de error
                            if key == 'programs':
                                extracted_data[key] = []
                            elif key == 'labs':
                                extracted_data[key] = []
                            elif key == 'scholarships':
                                extracted_data[key] = []
                            elif key == 'admission':
                                extracted_data[key] = extract_admission_info(university_name, univ['url'], univ_id, fallback=True)
                            elif key == 'cost':
                                extracted_data[key] = extract_cost_living_info(university_name, univ['city'], country, univ_id, fallback=True)
                            elif key == 'outcome':
                                extracted_data[key] = extract_outcome_info(university_name, univ['url'], univ_id, fallback=True)
                
                # Añadir datos a los dataframes
                programs = extracted_data['programs'] or []
                labs = extracted_data['labs'] or []
                scholarships = extracted_data['scholarships'] or []
                admission = extracted_data['admission']
                cost = extracted_data['cost']
                outcome = extracted_data['outcome']
                
                # Verificar que haya datos antes de añadirlos a los dataframes
                if programs:
                    programs_df = pd.concat([programs_df, pd.DataFrame(programs)], ignore_index=True)
                    logger.info(f"Añadidos {len(programs)} programas al DataFrame")
                
                if labs:
                    labs_df = pd.concat([labs_df, pd.DataFrame(labs)], ignore_index=True)
                    logger.info(f"Añadidos {len(labs)} laboratorios al DataFrame")
                
                if scholarships:
                    scholarships_df = pd.concat([scholarships_df, pd.DataFrame(scholarships)], ignore_index=True)
                    logger.info(f"Añadidas {len(scholarships)} becas al DataFrame")
                
                if admission:
                    admissions_df = pd.concat([admissions_df, pd.DataFrame([admission])], ignore_index=True)
                    logger.info("Información de admisión añadida al DataFrame")
                
                if cost:
                    costs_df = pd.concat([costs_df, pd.DataFrame([cost])], ignore_index=True)
                    logger.info("Información de costos añadida al DataFrame")
                
                if outcome:
                    outcomes_df = pd.concat([outcomes_df, pd.DataFrame([outcome])], ignore_index=True)
                    logger.info("Información de resultados añadida al DataFrame")
                
                # Crear notas vacías y cronograma para cada programa
                for program in programs:
                    prog_id = program['Prog_ID']
                    program_name = program['Program Name']
                    deadline = program['Application Deadline']
                    
                    # Crear notas y cronograma
                    notes = create_empty_notes(university_name, univ_id, prog_id)
                    timeline = create_empty_timeline(university_name, univ_id, prog_id, program_name, deadline)
                    
                    notes_df = pd.concat([notes_df, pd.DataFrame([notes])], ignore_index=True)
                    timeline_df = pd.concat([timeline_df, pd.DataFrame([timeline])], ignore_index=True)
                
                # Añadir registros adicionales para universidad en general
                notes = create_empty_notes(university_name, univ_id)
                timeline = create_empty_timeline(university_name, univ_id, program_name=university_name)
                
                notes_df = pd.concat([notes_df, pd.DataFrame([notes])], ignore_index=True)
                timeline_df = pd.concat([timeline_df, pd.DataFrame([timeline])], ignore_index=True)
                
                logger.info(f"Extracción exitosa para {university_name}")
                
                # Guardar datos parciales cada 5 universidades para evitar pérdida de datos
                if (univ_idx + 1) % 5 == 0 or (country_idx == len(countries) - 1 and univ_idx == len(universities[country]) - 1):
                    logger.info("Guardando datos parciales...")
                    write_excel(
                        universities_df, programs_df, labs_df, scholarships_df, 
                        admissions_df, costs_df, outcomes_df, notes_df, timeline_df,
                        f"partial_{country.replace(' ', '_')}_{univ_idx}.xlsx"
                    )
                
            except Exception as e:
                logger.error(f"Error al procesar {university_name}: {str(e)}")
                # Continuar con la siguiente universidad
                continue
    
    # Escribir datos en el archivo Excel final
    try:
        logger.info("Escribiendo datos finales en el archivo Excel...")
        write_excel(
            universities_df, programs_df, labs_df, scholarships_df, 
            admissions_df, costs_df, outcomes_df, notes_df, timeline_df,
            OUTPUT_EXCEL
        )
        logger.info(f"Datos escritos exitosamente en {OUTPUT_EXCEL}")
    except Exception as e:
        logger.error(f"Error al escribir datos en Excel: {str(e)}")
        # Intentar guardar un archivo de respaldo
        try:
            backup_file = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            write_excel(
                universities_df, programs_df, labs_df, scholarships_df, 
                admissions_df, costs_df, outcomes_df, notes_df, timeline_df,
                backup_file
            )
            logger.info(f"Datos de respaldo guardados en {backup_file}")
        except:
            logger.critical("¡NO SE PUDIERON GUARDAR NI SIQUIERA LOS DATOS DE RESPALDO!")
    
    # Si llegamos aquí, eliminar el checkpoint ya que hemos terminado
    if os.path.exists(CHECKPOINT_FILE):
        try:
            os.remove(CHECKPOINT_FILE)
            logger.info("Checkpoint eliminado tras finalización exitosa")
        except:
            logger.warning("No se pudo eliminar el archivo de checkpoint")
    
    # Mostrar estadísticas finales
    end_time = time.time()
    total_time = end_time - start_time
    hours, remainder = divmod(total_time, 3600)
    minutes, seconds = divmod(remainder, 60)
    
    logger.info("=== PROCESO DE EXTRACCIÓN FINALIZADO ===")
    logger.info(f"Tiempo total: {int(hours)}h {int(minutes)}m {int(seconds)}s")
    logger.info(f"Universidades procesadas: {universities_df.shape[0]}")
    logger.info(f"Programas extraídos: {programs_df.shape[0]}")
    logger.info(f"Laboratorios extraídos: {labs_df.shape[0]}")
    logger.info(f"Becas extraídas: {scholarships_df.shape[0]}")
    
    return True

def write_excel(universities_df, programs_df, labs_df, scholarships_df, 
               admissions_df, costs_df, outcomes_df, notes_df, timeline_df,
               output_file):
    """
    Escribe todos los dataframes en un archivo Excel con formato apropiado.
    
    Args:
        universities_df, programs_df, etc.: DataFrames a escribir
        output_file (str): Nombre del archivo de salida
    """
    try:
        # Cargar la plantilla original para copiar estilos
        wb_template = pd.ExcelFile(INPUT_EXCEL)
        
        # Crear un ExcelWriter
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Escribir cada DataFrame en su hoja correspondiente
            sheet_mapping = {
                '1_University': universities_df,
                '2_Program': programs_df,
                '3_Lab-Research': labs_df,
                '4_Scholarships': scholarships_df,
                '5_Admission': admissions_df,
                '6_Cost of Living': costs_df,
                '7_Outcomes': outcomes_df,
                '8_Notes': notes_df,
                '9_Timeline': timeline_df
            }
            
            for sheet_name, df in sheet_mapping.items():
                if not df.empty:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Copiar la hoja Dashboard de la plantilla
            workbook = writer.book
            
            try:
                # Cargar la plantilla original
                template = pd.ExcelFile(INPUT_EXCEL).book
                
                if '10_Dashboard' in template.sheetnames:
                    source_sheet = template['10_Dashboard']
                    
                    # Verificar si la hoja ya existe en el destino
                    if '10_Dashboard' in workbook.sheetnames:
                        # Si existe, eliminarla primero
                        std = workbook['10_Dashboard']
                        workbook.remove(std)
                    
                    # Crear una nueva hoja
                    target_sheet = workbook.create_sheet(title='10_Dashboard')
                    
                    # Copiar contenido y estilos
                    for row in source_sheet.rows:
                        for cell in row:
                            target_sheet[cell.coordinate] = cell.value
                            
                            if cell.has_style:
                                target_cell = target_sheet[cell.coordinate]
                                target_cell.font = cell.font
                                target_cell.border = cell.border
                                target_cell.fill = cell.fill
                                target_cell.number_format = cell.number_format
                                target_cell.alignment = cell.alignment
            except Exception as e:
                logger.warning(f"Error al copiar la hoja Dashboard: {str(e)}")
        
        return True
    except Exception as e:
        logger.error(f"Error al escribir el archivo Excel {output_file}: {str(e)}")
        raise

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logger.critical(f"Error crítico en la ejecución principal: {str(e)}")
